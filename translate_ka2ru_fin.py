import os
import re
import json
import time
import zipfile
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from typing import Dict, List, Set, Iterable, Callable, Optional, Any
from openai import RateLimitError

import xml.etree.ElementTree as ET

from transformers import AutoTokenizer, AutoModelForSeq2SeqLM, AutoModelForCausalLM
import torch

# ============ Разбиваем большой текст по токенам, чтобы не превышать ограничения ChatGPT =============

def estimate_tokens(text: str) -> int:
    """
    Консервативная оценка количества токенов.

    Раньше мы брали len(text)//3 и СИЛЬНО недооценивали.
    Теперь считаем примерно как количество символов, с небольшим запасом.
    Это заведомо завышает токены, но безопасно для лимита.
    """
    if not text:
        return 1
    # +20% запас сверху
    return int(len(text) * 1.2)


def split_fragments_by_tokens(
    fragments: List[Dict[str, Any]],
    max_tokens_per_batch: int = 8000,
) -> Iterable[List[Dict[str, Any]]]:
    """
    Делит список фрагментов на батчи так, чтобы суммарное
    оценочное кол-во токенов в одном батче не превышало max_tokens_per_batch.
    """
    batch: List[Dict[str, Any]] = []
    current_tokens = 0

    for frag in fragments:
        text = frag["text"]
        t = estimate_tokens(text)

        # Если фрагмент сам больше лимита, отправляем его отдельно
        if t > max_tokens_per_batch:
            if batch:
                yield batch
                batch = []
                current_tokens = 0
            yield [frag]
            continue

        # Если добавление фрагмента переполнит партию – отдаем текущую
        if batch and current_tokens + t > max_tokens_per_batch:
            yield batch
            batch = []
            current_tokens = 0

        batch.append(frag)
        current_tokens += t

    if batch:
        yield batch

def split_texts_by_tokens(
    texts: List[str],
    max_tokens_per_batch: int = 8000,
) -> Iterable[List[str]]:
    """
    Делит список строк на батчи так, чтобы суммарное
    оценочное число токенов в батче не превышало max_tokens_per_batch.
    Использует estimate_tokens(), как и для перевода.
    """
    batch: List[str] = []
    current_tokens = 0

    for text in texts:
        if not isinstance(text, str):
            continue
        t = estimate_tokens(text)

        # если один текст сам по себе больше лимита — отправляем его в отдельном батче
        if t > max_tokens_per_batch:
            if batch:
                yield batch
                batch = []
                current_tokens = 0
            yield [text]
            continue

        # если добавление этого текста переполнит батч — отдаем текущий и начинаем новый
        if batch and current_tokens + t > max_tokens_per_batch:
            yield batch
            batch = []
            current_tokens = 0

        batch.append(text)
        current_tokens += t

    if batch:
        yield batch


# ============ Настройки по умолчанию ============

DEFAULT_CHATGPT_MODEL = "gpt-4.1-mini"
CHATGPT_MODELS = [
    "gpt-4.1-mini",
    "gpt-4.1",
    "gpt-4o-mini",
    "gpt-4o",
    "gpt-5.1",
    "gpt-5-mini",
]

# направления перевода: грузинский -> целевой язык
DIRECTION_CONFIG: Dict[str, Dict[str, str]] = {
    "ka-ru": {
        "label": "Грузинский → Русский",
        "target_language": "Russian",
        "suffix": "_ru",
    },
    "ka-en": {
        "label": "Грузинский → Английский",
        "target_language": "English",
        "suffix": "_en",
    },
}


# диапазоны Unicode для грузинского
GEORGIAN_RE = re.compile(r"[\u10A0-\u10FF\u1C90-\u1CBF]+")


# ============ Утилиты ============

def is_docx(path: str) -> bool:
    return path.lower().endswith(".docx")


def is_xlsx(path: str) -> bool:
    return path.lower().endswith(".xlsx")


def chunks(lst: List[str], n: int) -> Iterable[List[str]]:
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def load_api_key_from_env_file(env_path: str) -> str:
    """
    Читает .env-файл и ищет OPENAI_API_KEY / API_KEY.
    Если не нашёл — берёт первую непустую, некомментированную строку как ключ.
    """
    if not os.path.exists(env_path):
        raise FileNotFoundError(f".env файл не найден: {env_path}")

    candidate = None
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip()
                if k in ("OPENAI_API_KEY", "API_KEY"):
                    return v
                if candidate is None and v:
                    candidate = v
            else:
                if candidate is None:
                    candidate = line

    if not candidate:
        raise ValueError("Не удалось извлечь ключ API из .env файла.")
    return candidate


def get_direction_code_from_label(label: str) -> str:
    for code, meta in DIRECTION_CONFIG.items():
        if meta["label"] == label:
            return code
    raise ValueError(f"Неизвестное направление перевода: {label}")


# ============ Работа с XML внутри DOCX/XLSX ============

def collect_docx_items(path: str) -> List[Dict[str, object]]:
    """
    DOCX — работаем по АБЗАЦАМ (<w:p>), но теперь собираем не просто тексты,
    а ПОЛНЫЙ список элементов с позиционными ID.

    Возвращает список словарей:
      {
        "id": "word/document.xml::p17",
        "xml_name": "word/document.xml",
        "p_index": 17,
        "full_text": "<весь текст абзаца как есть>",
        "clean_text": "<full_text.strip()>"
      }

    Собираем только те абзацы, в которых есть грузинский текст.
    """
    items: List[Dict[str, object]] = []

    with zipfile.ZipFile(path, "r") as zin:
        for info in zin.infolist():
            fname = info.filename
            if not (fname.startswith("word/") and fname.lower().endswith(".xml")):
                continue

            xml_bytes = zin.read(fname)
            try:
                root = ET.fromstring(xml_bytes)
            except Exception:
                continue

            m = re.match(r"\{(.*)\}", root.tag)
            ns = m.group(1) if m else ""
            p_tag = f"{{{ns}}}p"
            t_tag = f"{{{ns}}}t"

            for p_index, p in enumerate(root.iter(p_tag)):
                t_elems = list(p.iter(t_tag))
                if not t_elems:
                    continue

                parts = [(t.text or "") for t in t_elems]
                full_text = "".join(parts)
                if not full_text:
                    continue

                if not GEORGIAN_RE.search(full_text):
                    continue

                clean_text = full_text.strip()
                if not clean_text:
                    continue

                item_id = f"{fname}::p{p_index}"
                items.append({
                    "id": item_id,
                    "xml_name": fname,
                    "p_index": p_index,
                    "full_text": full_text,
                    "clean_text": clean_text,
                })

    print(f"📄 DOCX: найдено {len(items)} абзацев с грузинским текстом.")
    return items


def collect_georgian_fragments_from_xml_bytes(xml_bytes: bytes) -> Set[str]:
    """
    Универсальная функция: находим все текстовые узлы, содержащие грузинский,
    и берём их полный текст целиком (strip()).
    """
    result: Set[str] = set()
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return result

    for elem in root.iter():
        text = elem.text
        if not text:
            continue
        if GEORGIAN_RE.search(text):
            cleaned = text.strip()
            if cleaned:
                result.add(cleaned)
    return result


def collect_fragments_xlsx(path: str) -> Set[str]:
    """
    XLSX — обрабатываем sharedStrings + worksheets.
    workbook.xml (имена листов) не трогаем.
    """
    to_translate: Set[str] = set()

    with zipfile.ZipFile(path, "r") as zin:
        for info in zin.infolist():
            fname = info.filename.lower()

            if fname.startswith("xl/sharedstrings") and fname.endswith(".xml"):
                xml_bytes = zin.read(info.filename)
            elif fname.startswith("xl/worksheets/") and fname.endswith(".xml"):
                xml_bytes = zin.read(info.filename)
            else:
                continue

            frags = collect_georgian_fragments_from_xml_bytes(xml_bytes)
            to_translate.update(frags)

    print(f"📊 XLSX: найдено {len(to_translate)} уникальных грузинских фрагментов.")
    return to_translate


def replace_georgian_in_xml_bytes(xml_bytes: bytes, mapping: Dict[str, str]) -> bytes:
    """
    Для XLSX: если текст узла (strip) есть в mapping — заменяем целиком,
    сохраняя ведущие/хвостовые пробелы.
    """
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return xml_bytes

    for elem in root.iter():
        text = elem.text
        if not text:
            continue

        stripped = text.strip()
        if stripped in mapping:
            prefix_len = len(text) - len(text.lstrip())
            suffix_len = len(text) - len(text.rstrip())
            prefix = text[:prefix_len]
            suffix = text[len(text) - suffix_len:] if suffix_len > 0 else ""
            new_text = mapping[stripped]
            elem.text = f"{prefix}{new_text}{suffix}"

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def process_docx_xml_paragraphs(
    xml_bytes: bytes,
    xml_name: str,
    id_mapping: Dict[str, str],
) -> bytes:
    """
    Пробегаем по <w:p> в одном XML-файле DOCX и, если для параграфа есть
    перевод в id_mapping, подменяем его текст. При этом:
      - если в текущем абзаце уже НЕТ грузинских букв, мы его НЕ трогаем,
        даже если его ID есть в id_mapping (защита от повторного прогона
        по уже переведённому файлу / вручную переведённым кускам).
    """
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return xml_bytes

    m = re.match(r"\{(.*)\}", root.tag)
    ns = m.group(1) if m else ""
    p_tag = f"{{{ns}}}p"
    t_tag = f"{{{ns}}}t"

    for p_index, p in enumerate(root.iter(p_tag)):
        para_id = f"{xml_name}::p{p_index}"
        if para_id not in id_mapping:
            continue

        t_elems = list(p.iter(t_tag))
        if not t_elems:
            continue

        orig_parts = [(t.text or "") for t in t_elems]
        orig_full = "".join(orig_parts)
        if not orig_full:
            continue

        # если в абзаце уже нет грузинских букв — не трогаем
        if not GEORGIAN_RE.search(orig_full):
            continue

        translated_clean = id_mapping[para_id]

        # сохраняем ведущие/хвостовые пробелы абзаца
        lead = len(orig_full) - len(orig_full.lstrip())
        trail = len(orig_full) - len(orig_full.rstrip())
        prefix = orig_full[:lead]
        suffix = orig_full[len(orig_full) - trail:] if trail > 0 else ""
        translated_full = prefix + translated_clean + suffix

        # Весь текст абзаца кладём в первый <w:t>, остальные чистим.
        t_elems[0].text = translated_full
        for t in t_elems[1:]:
            t.text = ""

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def debug_scan_docx_for_georgian(path: str, max_examples: int = 20) -> None:
    """
    Отладочная функция: сканирует DOCX и ищет все абзацы, где остался грузинский текст.
    Печатает количество и несколько примеров (кусочек текста + имя XML-файла + индекс абзаца).
    """
    count = 0
    examples = []

    with zipfile.ZipFile(path, "r") as zin:
        for info in zin.infolist():
            fname = info.filename
            if not (fname.startswith("word/") and fname.lower().endswith(".xml")):
                continue

            xml_bytes = zin.read(fname)
            try:
                root = ET.fromstring(xml_bytes)
            except Exception:
                continue

            m = re.match(r"\{(.*)\}", root.tag)
            ns = m.group(1) if m else ""
            p_tag = f"{{{ns}}}p"
            t_tag = f"{{{ns}}}t"

            for p_index, p in enumerate(root.iter(p_tag)):
                parts = []
                for t in p.iter(t_tag):
                    parts.append(t.text or "")
                full_text = "".join(parts)
                if not full_text:
                    continue
                if GEORGIAN_RE.search(full_text):
                    count += 1
                    if len(examples) < max_examples:
                        snippet = full_text.strip()
                        if len(snippet) > 120:
                            snippet = snippet[:117] + "..."
                        examples.append((fname, p_index, snippet))

    print(f"🔍 В файле {os.path.basename(path)} осталось абзацев с грузинским: {count}")
    for fname, p_index, snippet in examples:
        print(f"  - {fname}::p{p_index}: {snippet}")


def apply_translations_docx(
    input_path: str,
    output_path: str,
    id_mapping: Dict[str, str],
    text_mapping: Optional[Dict[str, str]] = None,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 90.0,
    end: float = 100.0,
) -> None:
    with zipfile.ZipFile(input_path, "r") as zin, \
         zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:

        infos = zin.infolist()
        total = len(infos)
        changed = 0

        for idx, info in enumerate(infos, start=1):
            fname = info.filename
            data = zin.read(fname)

            new_data = data

            # 1) для word/*.xml прогоняем абзацы с id_mapping
            if fname.startswith("word/") and fname.lower().endswith(".xml"):
                new_data = process_docx_xml_paragraphs(new_data, fname, id_mapping)

            # 2) для ЛЮБОГО *.xml — общая замена по text_mapping
            if fname.lower().endswith(".xml") and text_mapping:
                new_data = replace_georgian_in_xml_bytes(new_data, text_mapping)

            if new_data != data:
                changed += 1

            zout.writestr(info, new_data)

            if progress_callback and total > 0:
                frac = idx / total
                pct = start + (end - start) * frac
                progress_callback(pct, "Применение перевода в DOCX...")

    print(f"💾 DOCX сохранён: {output_path}, изменённых XML: {changed}")


def apply_translations_xlsx(
    input_path: str,
    output_path: str,
    text_mapping: Dict[str, str],
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 90.0,
    end: float = 100.0,
) -> None:
    """
    Безопасно применяет переводы к XLSX через openpyxl.
    """
    from openpyxl import load_workbook

    wb = load_workbook(input_path, data_only=False)

    total_cells = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            total_cells += len(row)
    if total_cells == 0:
        total_cells = 1

    processed = 0
    changed_cells = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str):
                    original = val
                    stripped = val.strip()
                    if stripped in text_mapping:
                        new_core = text_mapping[stripped]

                        prefix_len = len(original) - len(original.lstrip())
                        suffix_len = len(original) - len(original.rstrip())
                        prefix = original[:prefix_len]
                        suffix = original[len(original) - suffix_len:] if suffix_len > 0 else ""

                        cell.value = f"{prefix}{new_core}{suffix}"
                        changed_cells += 1

                processed += 1
                if progress_callback:
                    frac = processed / total_cells
                    pct = start + (end - start) * frac
                    progress_callback(pct, "Применение перевода в XLSX...")

    wb.save(output_path)
    print(f"💾 XLSX сохранён: {output_path}, изменённых ячеек: {changed_cells}")


# ============ Переводчики (ChatGPT) ============

def translate_with_chatgpt(
    fragments: List[str],
    model_name: str,
    api_key: str,
    target_language: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 10.0,
    end: float = 60.0,
) -> Dict[str, str]:
    """
    Перевод грузинских фрагментов через ChatGPT (OpenAI API),
    с разбиением по токенам и обработкой RateLimitError.
    """
    from openai import OpenAI

    os.environ["OPENAI_API_KEY"] = api_key
    client = OpenAI()

    # Чистим и выкидываем пустое
    cleaned = []
    for f in fragments:
        s = (f or "").strip()
        if s:
            cleaned.append(s)

    # Уникализируем
    unique_texts: List[str] = []
    seen: Set[str] = set()
    for s in cleaned:
        if s not in seen:
            seen.add(s)
            unique_texts.append(s)

    if not unique_texts:
        return {}

    # Назначаем ID
    id_to_text: Dict[int, str] = {i: txt for i, txt in enumerate(unique_texts)}
    text_to_id: Dict[str, int] = {txt: i for i, txt in id_to_text.items()}

    print(f"Для перевода через ChatGPT подготовлено {len(unique_texts)} уникальных фрагментов.")

    # Готовим фрагменты в формате {id, text} для токенного батчинга
    fragments_struct: List[Dict[str, Any]] = [
        {"id": i, "text": txt}
        for i, txt in id_to_text.items()
    ]

    # Разбиваем по токенам (очень консервативно)
    batches = list(split_fragments_by_tokens(fragments_struct, max_tokens_per_batch=8000))
    print(f"Будет отправлено {len(batches)} батч(ей) в модель {model_name}.")

    total = len(unique_texts)
    done = 0
    id_to_translated: Dict[int, str] = {}

    for batch_idx, batch in enumerate(batches, start=1):
        batch_items = batch  # каждый элемент: {"id": int, "text": str}

        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Перевод через ChatGPT...")

        user_payload = {
            "source_language": "Georgian",
            "target_language": target_language,
            "items": batch_items,
        }

        # ТВОЙ system_msg — без изменений
        system_msg = (
            "You are a professional legal and technical translator. "
            "The texts are official documents (tariff methodology, regulatory acts, explanatory notes). "
            f"Translate from Georgian to {target_language} into natural, formal, human-quality {target_language}. "
            "You MAY freely change word order, grammar, and morphology so that the result sounds like good native legal language, "
            "but you MUST preserve all facts, numbers, names, and logical relations. "
            "Avoid literal calques from Georgian where they sound unnatural in the target language. "
            "Do NOT merge separate words together: always keep proper spaces between words, "
            "between prepositions and nouns, and around conjunctions. "
            "Fix any missing spaces if they are present in the original. "
            "Return ONLY a JSON object with a single key 'translations', whose value is a list of objects "
            "of the form {\"id\": <same id>, \"text\": <translation>}. "
            "Do not add extra fields."
        )

        # ==== вызов модели с ретраями на RateLimit ====
        max_retries = 5
        delay_seconds = 10

        for attempt in range(1, max_retries + 1):
            try:
                resp = client.chat.completions.create(
                    model=model_name,
                    response_format={"type": "json_object"},
                    messages=[
                        {"role": "system", "content": system_msg},
                        {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
                    ],
                )
                break
            except RateLimitError:
                print(
                    f"[Batch {batch_idx}/{len(batches)}] "
                    f"Перевышен лимит токенов/минуту (попытка {attempt}/{max_retries}). "
                    f"Ждём {delay_seconds} секунд..."
                )
                if attempt == max_retries:
                    raise
                time.sleep(delay_seconds)
        else:
            # Теоретически сюда не дойдём (выше raise), но оставим на всякий
            raise RuntimeError("Не удалось получить ответ от ChatGPT после нескольких попыток.")

        content = resp.choices[0].message.content
        try:
            data = json.loads(content)
        except json.JSONDecodeError:
            print("⚠️ Ошибка JSON от модели, ответ:")
            print(content)
            raise

        translations_list = data.get("translations")
        if not isinstance(translations_list, list):
            raise ValueError("Ожидался ключ 'translations' со списком объектов {id, text}.")

        for obj in translations_list:
            if not isinstance(obj, dict):
                continue
            tid = obj.get("id")
            ttext = obj.get("text")
            try:
                tid_int = int(tid)
            except (TypeError, ValueError):
                continue
            if not isinstance(ttext, str) or not ttext.strip():
                continue
            id_to_translated[tid_int] = ttext

        done += len(batch_items)
        print(f"   ChatGPT перевёл {done}/{total} уникальных фрагментов (batch {batch_idx}/{len(batches)})")

        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Перевод через ChatGPT...")

    # Собираем mapping: исходный текст -> перевод
    mapping: Dict[str, str] = {}
    for txt, tid in text_to_id.items():
        trans = id_to_translated.get(tid)
        if isinstance(trans, str) and trans.strip():
            mapping[txt] = trans
        else:
            mapping[txt] = txt

    return mapping

def post_edit_with_chatgpt(
    mapping: Dict[str, str],
    model_name: str,
    api_key: str,
    target_language: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 70.0,
    end: float = 90.0,
) -> Dict[str, str]:
    """
    Литературная вычитка уже переведённого текста через ChatGPT,
    с аккуратным батчингом по токенам и обработкой RateLimitError.
    """
    from openai import OpenAI

    os.environ["OPENAI_API_KEY"] = api_key
    client = OpenAI()

    unique_values: List[str] = []
    seen = set()
    for v in mapping.values():
        if isinstance(v, str) and v.strip() and v not in seen:
            seen.add(v)
            unique_values.append(v)

    if not unique_values:
        return mapping

    total = len(unique_values)
    done = 0

    # делим не по количеству штук, а по оценочным токенам
    batches = list(split_texts_by_tokens(unique_values, max_tokens_per_batch=8000))
    print(f"[Post-edit] Будет отправлено {len(batches)} батч(ей) в модель {model_name}.")

    improved_map: Dict[str, str] = {}

    system_msg = (
        "You are a professional editor for legal, regulatory and technical documents. "
        f"Improve style, clarity, grammar and fluency in {target_language} while preserving the same meaning, facts, numbers and legal content. "
        "You MAY change word order, fix awkward literal phrases, adjust cases and morphology, "
        "replace unnatural calques with standard legal expressions, and break or merge sentences if it improves readability. "
        "Do NOT add new facts or remove existing ones. "
        "Return ONLY a JSON object mapping each original text to its improved version. "
        "Keys MUST be EXACTLY the original texts. Do not add extra fields."
    )

    max_retries = 5
    delay_seconds = 10

    for batch_idx, batch in enumerate(batches, start=1):
        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Литературная вычитка перевода (ChatGPT)...")

        user_payload = {
            "target_language": target_language,
            "texts": batch,
        }

        for attempt in range(1, max_retries + 1):
            try:
                resp = client.chat.completions.create(
                    model=model_name,
                    response_format={"type": "json_object"},
                    messages=[
                        {"role": "system", "content": system_msg},
                        {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
                    ],
                )
                break
            except RateLimitError as e:
                msg = str(e)
                print(
                    f"[Post-edit batch {batch_idx}/{len(batches)}] "
                    f"Перевышен лимит (попытка {attempt}/{max_retries}): {msg}"
                )
                # если ошибка говорит, что запрос слишком большой, ретраи не помогут
                if "Request too large" in msg and "tokens per min" in msg:
                    raise
                if attempt == max_retries:
                    raise
                time.sleep(delay_seconds)
        else:
            raise RuntimeError("Не удалось получить ответ от ChatGPT (post_edit) после нескольких попыток.")

        content = resp.choices[0].message.content
        try:
            data = json.loads(content)
        except json.JSONDecodeError:
            print("⚠️ Ошибка JSON от модели (вычитка), ответ:")
            print(content)
            raise

        for orig_text in batch:
            new_text = data.get(orig_text)
            if isinstance(new_text, str) and new_text.strip():
                improved_map[orig_text] = new_text
            else:
                improved_map[orig_text] = orig_text

        done += len(batch)
        print(f"   ChatGPT вычитал {done}/{total} фрагментов (batch {batch_idx}/{len(batches)})")

        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Литературная вычитка перевода (ChatGPT)...")

    # Собираем новый mapping: грузинский -> улучшенный перевод
    new_mapping: Dict[str, str] = {}
    for geo, raw_trans in mapping.items():
        if isinstance(raw_trans, str):
            new_mapping[geo] = improved_map.get(raw_trans, raw_trans)
        else:
            new_mapping[geo] = raw_trans

    return new_mapping



def fix_spacing_with_chatgpt(
    mapping: Dict[str, str],
    model_name: str,
    api_key: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 70.0,
    end: float = 90.0,
) -> Dict[str, str]:
    """
    Аккуратная правка ПРОБЕЛОВ в уже переведённом русском тексте,
    с батчингом по токенам и обработкой RateLimitError.
    """
    from openai import OpenAI

    os.environ["OPENAI_API_KEY"] = api_key
    client = OpenAI()

    unique_values: List[str] = []
    seen = set()
    for v in mapping.values():
        if isinstance(v, str) and v.strip() and v not in seen:
            seen.add(v)
            unique_values.append(v)

    if not unique_values:
        return mapping

    total = len(unique_values)
    done = 0

    batches = list(split_texts_by_tokens(unique_values, max_tokens_per_batch=8000))
    print(f"[Fix-spacing] Будет отправлено {len(batches)} батч(ей) в модель {model_name}.")

    fixed_map: Dict[str, str] = {}

    system_msg = (
        "You receive Russian texts which are already translated correctly. "
        "Your ONLY task is to fix spacing errors: insert or delete ASCII space characters (U+0020) "
        "where necessary between words, numbers and punctuation, and collapse multiple spaces to single ones if appropriate. "
        "You MUST NOT change, delete, reorder or insert ANY non-space characters (letters, digits, punctuation). "
        "Return ONLY a JSON object mapping each original string to its corrected version. "
        "Keys MUST be EXACTLY the original strings. Do not add extra fields."
    )

    max_retries = 5
    delay_seconds = 10

    for batch_idx, batch in enumerate(batches, start=1):
        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Правка пробелов в русском тексте...")

        user_payload = {
            "texts": batch,
        }

        for attempt in range(1, max_retries + 1):
            try:
                resp = client.chat.completions.create(
                    model=model_name,
                    response_format={"type": "json_object"},
                    messages=[
                        {"role": "system", "content": system_msg},
                        {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
                    ],
                )
                break
            except RateLimitError as e:
                msg = str(e)
                print(
                    f"[Fix-spacing batch {batch_idx}/{len(batches)}] "
                    f"Перевышен лимит (попытка {attempt}/{max_retries}): {msg}"
                )
                if "Request too large" in msg and "tokens per min" in msg:
                    raise
                if attempt == max_retries:
                    raise
                time.sleep(delay_seconds)
        else:
            raise RuntimeError("Не удалось получить ответ от ChatGPT (fix_spacing) после нескольких попыток.")

        content = resp.choices[0].message.content
        try:
            data = json.loads(content)
        except json.JSONDecodeError:
            print("⚠️ Ошибка JSON от модели (fix_spacing), ответ:")
            print(content)
            raise

        for orig_text in batch:
            new_text = data.get(orig_text)
            if isinstance(new_text, str) and new_text.strip():
                fixed_map[orig_text] = new_text
            else:
                fixed_map[orig_text] = orig_text

        done += len(batch)
        print(f"   ChatGPT поправил пробелы в {done}/{total} фрагментах (batch {batch_idx}/{len(batches)})")

        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Правка пробелов в русском тексте...")

    new_mapping: Dict[str, str] = {}
    for geo, ru in mapping.items():
        if isinstance(ru, str):
            new_mapping[geo] = fixed_map.get(ru, ru)
        else:
            new_mapping[geo] = ru

    return new_mapping



# ============ Переводчик NLLB (локальный) ============

def translate_with_local_model(
    fragments: List[str],
    direction_code: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 10.0,
    end: float = 90.0,
) -> Dict[str, str]:
    """
    Универсальный ЛОКАЛЬНЫЙ переводчик на основе NLLB-200.
    """

    LANG_MAP = {
        "ka": "kat_Geor",
        "ru": "rus_Cyrl",
        "en": "eng_Latn",
    }

    if "-" not in direction_code:
        raise ValueError(f"direction_code должен быть формата ka-ru, а получено: {direction_code}")

    src, tgt = direction_code.split("-")

    if src not in LANG_MAP:
        raise ValueError(f"Источник языка '{src}' не поддержан в LANG_MAP")

    if tgt not in LANG_MAP:
        raise ValueError(f"Целевой язык '{tgt}' не поддержан в LANG_MAP")

    SRC_LANG = LANG_MAP[src]
    TGT_LANG = LANG_MAP[tgt]

    remaining = [f.strip() for f in fragments if isinstance(f, str) and f.strip()]
    if not remaining:
        return {}

    MODEL_NAME = "facebook/nllb-200-3.3B"

    print(f"⏳ Загружаем локальную NLLB модель: {MODEL_NAME}")
    if progress_callback:
        progress_callback(start, f"Загрузка локальной модели NLLB ({MODEL_NAME})…")

    tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME, src_lang=SRC_LANG)
    device = torch.device("cpu")
    model = AutoModelForSeq2SeqLM.from_pretrained(MODEL_NAME).to(device)
    model.eval()

    BATCH_SIZE = 1
    MAX_TOKENS = 256

    total = len(remaining)
    done = 0
    mapping: Dict[str, str] = {}

    for i, batch in enumerate(chunks(remaining, BATCH_SIZE), start=1):
        print(f"--> [NLLB {direction_code}] batch {i}, size={len(batch)}")

        inputs = tokenizer(
            batch,
            return_tensors="pt",
            padding=True,
            truncation=True,
            max_length=512,
        )
        inputs = {k: v.to(device) for k, v in inputs.items()}

        bos_id = tokenizer.convert_tokens_to_ids(TGT_LANG)
        if bos_id is None:
            raise RuntimeError(f"Не удалось получить token id для языка {TGT_LANG}")

        with torch.no_grad():
            generated = model.generate(
                **inputs,
                forced_bos_token_id=bos_id,
                max_length=MAX_TOKENS,
            )

        outputs = tokenizer.batch_decode(generated, skip_special_tokens=True)

        for orig, trans in zip(batch, outputs):
            trans = (trans or "").strip()
            mapping[orig] = trans if trans else orig

        done += len(batch)
        if progress_callback and total > 0:
            pct = start + (end - start) * (done / total)
            progress_callback(pct, f"Перевод локальной моделью (NLLB, {direction_code})…")

        print(f"   [NLLB {direction_code}] готово {done}/{total}")

    return mapping


# ============ Локальная литературная вычитка (Qwen) ============

def post_edit_with_qwen_local(
    mapping: Dict[str, str],
    target_language: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 60.0,
    end: float = 90.0,
) -> Dict[str, str]:
    """
    Локальная литературная вычитка перевода с помощью Qwen2.5-3B-Instruct.

    mapping: {грузинский_оригинал -> машинный_перевод (ru/en)}
    target_language: "Russian" / "English" (из DIRECTION_CONFIG)
    """
    MODEL_NAME = "Qwen/Qwen2.5-3B-Instruct"

    # Собираем уникальные значения
    unique_values: List[str] = []
    seen = set()
    for v in mapping.values():
        if isinstance(v, str) and v.strip() and v not in seen:
            seen.add(v)
            unique_values.append(v)

    if not unique_values:
        return mapping

    print(f"⏳ Загружаем локальную модель вычитки: {MODEL_NAME}")
    if progress_callback:
        progress_callback(start, f"Загрузка Qwen2.5-3B-Instruct для вычитки…")

    tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    if device.type == "cuda":
        model = AutoModelForCausalLM.from_pretrained(
            MODEL_NAME,
            torch_dtype=torch.float16,
            device_map="auto",
        )
    else:
        model = AutoModelForCausalLM.from_pretrained(MODEL_NAME).to(device)
    model.eval()

    system_msg_text = (
        f"You are a professional editor for {target_language} legal, regulatory and technical documents. "
        f"Improve style, clarity, grammar and fluency in {target_language} while preserving the same meaning, "
        "facts, numbers and legal content. "
        "You MAY change word order, fix awkward literal phrases, adjust cases and morphology, "
        "replace unnatural calques with standard legal expressions, and break or merge sentences if it improves readability. "
        "Do NOT add new facts or remove existing ones. "
        "Return ONLY the improved text, without explanations, without quotes."
    )

    BATCH_SIZE = 4
    total = len(unique_values)
    done = 0

    improved_map: Dict[str, str] = {}

    for batch in chunks(unique_values, BATCH_SIZE):
        for text in batch:
            # Chat-шаблон для Qwen
            messages = [
                {"role": "system", "content": system_msg_text},
                {"role": "user", "content": text},
            ]
            prompt = tokenizer.apply_chat_template(
                messages,
                tokenize=False,
                add_generation_prompt=True,
            )

            inputs = tokenizer(prompt, return_tensors="pt").to(device)

            with torch.no_grad():
                output_ids = model.generate(
                    **inputs,
                    max_new_tokens=512,
                    do_sample=False,
                )

            # Обрезаем префикс-промпт и оставляем только сгенерированный кусок
            gen_ids = output_ids[0][inputs["input_ids"].shape[1]:]
            out_text = tokenizer.decode(gen_ids, skip_special_tokens=True).strip()

            if not out_text:
                out_text = text

            improved_map[text] = out_text

        done += len(batch)
        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Литературная вычитка (локальная Qwen)…")

        print(f"   Qwen вычитал {done}/{total} фрагментов")

    # Собираем новый mapping: грузинский -> улучшенный перевод
    new_mapping: Dict[str, str] = {}
    for geo, raw_trans in mapping.items():
        if isinstance(raw_trans, str):
            new_mapping[geo] = improved_map.get(raw_trans, raw_trans)
        else:
            new_mapping[geo] = raw_trans

    return new_mapping


# ============ Простая правка пробелов (регэксы) ============

def normalize_segment_boundaries(segments: List[str]) -> List[str]:
    if len(segments) <= 1:
        return segments

    segs = segments[:]

    for i in range(len(segs) - 1):
        left = segs[i]
        right = segs[i + 1]

        left = re.sub(r' {2,}', ' ', left)
        right = re.sub(r' {2,}', ' ', right)

        if left.endswith(" ") and right.startswith(" "):
            right = right.lstrip()

        segs[i] = left
        segs[i + 1] = right

    return segs


def fix_basic_spacing_ru(text: str) -> str:
    import re

    text = text.replace("\u00A0", " ")

    text = re.sub(r'([А-ЯЁа-яё])№\s*(\d)', r'\1 №\2', text)
    text = re.sub(r'№\s*(\d)', r'№ \1', text)

    text = re.sub(r'(?i)\b(от|до|по|на|в|к|с|у)(\d)', r'\1 \2', text)

    text = re.sub(r'(\d)([А-ЯЁа-яё])', r'\1 \2', text)

    text = re.sub(r'(\d{3,4})\s*(год[ауе]?|гг?\.?)', r'\1 \2', text)

    text = re.sub(r',([^\s])', r', \1', text)
    text = re.sub(r';([^\s])', r'; \1', text)
    text = re.sub(r'([^.])\.([А-ЯЁ])', r'\1. \2', text)

    text = re.sub(r'(%)([А-ЯЁа-яё])', r'\1 \2', text)

    text = re.sub(r'[ \t]{2,}', ' ', text)

    return text


# ============ Логика перевода одного файла ============

def process_file(
    file_path: str,
    translator_kind: str,
    chatgpt_model: str,
    env_path: Optional[str],
    direction_code: str,
    post_edit: bool,
    progress_callback: Optional[Callable[[float, str], None]] = None,
) -> str:
    """
    Главная функция: собирает фрагменты, переводит выбранным движком,
    при необходимости делает литературную вычитку,
    применяет переводы, возвращает путь к выходному файлу.
    """

    if progress_callback is None:
        def progress_callback(pct: float, msg: str) -> None:
            pass  # заглушка

    if not (is_docx(file_path) or is_xlsx(file_path)):
        raise ValueError("Поддерживаются только файлы .docx и .xlsx")

    if direction_code not in DIRECTION_CONFIG:
        raise ValueError(f"Неизвестное направление: {direction_code}")

    meta = DIRECTION_CONFIG[direction_code]
    target_language = meta["target_language"]
    suffix = meta["suffix"]

    # 1. Сбор фрагментов
    progress_callback(0.0, "Сбор грузинских фрагментов...")

    if is_docx(file_path):
        items = collect_docx_items(file_path)
        if not items:
            progress_callback(0.0, "Грузинский текст не найден.")
            raise RuntimeError("В файле не найден грузинский текст для перевода.")

        base_texts = {str(it["clean_text"]) for it in items}

        extra_texts: Set[str] = set()
        with zipfile.ZipFile(file_path, "r") as zin:
            for info in zin.infolist():
                fname = info.filename
                if not fname.lower().endswith(".xml"):
                    continue

                xml_bytes = zin.read(fname)
                extra_texts.update(collect_georgian_fragments_from_xml_bytes(xml_bytes))

        all_texts = base_texts | extra_texts

        fragments_for_translation = sorted(
            t for t in all_texts
            if t.strip() and GEORGIAN_RE.search(t)
        )

        items_for_docx = items
    else:
        fragments_set = collect_fragments_xlsx(file_path)
        if not fragments_set:
            progress_callback(0.0, "Грузинский текст не найден.")
            raise RuntimeError("В файле не найден грузинский текст для перевода.")
        fragments_for_translation = sorted(fragments_set)
        items_for_docx = None

    print(f"Найдено {len(fragments_for_translation)} уникальных фрагментов для перевода.")
    progress_callback(5.0, f"Найдено {len(fragments_for_translation)} фрагментов. Подготовка к переводу...")

    # 2. Перевод
    if translator_kind == "chatgpt":
        if not env_path:
            raise ValueError("Не выбран .env файл с токеном для ChatGPT.")
        api_key = load_api_key_from_env_file(env_path)

        mapping_text_to_trans = translate_with_chatgpt(
            fragments_for_translation,
            chatgpt_model,
            api_key,
            target_language,
            progress_callback=progress_callback,
            start=10.0,
            end=60.0,
        )

        if post_edit:
            mapping_text_to_trans = post_edit_with_chatgpt(
                mapping_text_to_trans,
                chatgpt_model,
                api_key,
                target_language,
                progress_callback=progress_callback,
                start=60.0,
                end=90.0,
            )

    else:
        # Локальный перевод NLLB
        mapping_text_to_trans = translate_with_local_model(
            fragments_for_translation,
            direction_code,
            progress_callback=progress_callback,
            start=10.0,
            end=60.0,
        )

        # Локальная литературная вычитка Qwen (если включена)
        if post_edit:
            mapping_text_to_trans = post_edit_with_qwen_local(
                mapping_text_to_trans,
                target_language=target_language,
                progress_callback=progress_callback,
                start=60.0,
                end=90.0,
            )

    # 2b. Преобразуем маппинг для DOCX в формат id -> translation
    if is_docx(file_path):
        id_mapping: Dict[str, str] = {}
        for it in items_for_docx:  # type: ignore
            clean_text = str(it["clean_text"])
            item_id = str(it["id"])
            translated = mapping_text_to_trans.get(clean_text, clean_text)
            id_mapping[item_id] = translated
    else:
        id_mapping = mapping_text_to_trans

    # 3. Применяем переводы к файлу
    base, ext = os.path.splitext(file_path)
    output_path = f"{base}{suffix}{ext}"

    progress_callback(90.0, "Применяем переводы к файлу...")
    if is_docx(file_path):
        apply_translations_docx(
            file_path,
            output_path,
            id_mapping,
            mapping_text_to_trans,
            progress_callback=progress_callback,
            start=90.0,
            end=100.0,
        )
    else:
        apply_translations_xlsx(
            file_path,
            output_path,
            id_mapping,
            progress_callback=progress_callback,
            start=90.0,
            end=100.0,
        )

    progress_callback(100.0, "Готово.")
    if is_docx(output_path):
        debug_scan_docx_for_georgian(output_path)

    return output_path


# ============ GUI (Tkinter) ============

class TranslatorGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Перевод грузинского текста в DOCX/XLSX")

        self.file_path_var = tk.StringVar()
        self.env_path_var = tk.StringVar()
        self.translator_var = tk.StringVar(value="chatgpt")
        self.model_var = tk.StringVar(value=DEFAULT_CHATGPT_MODEL)
        self.direction_label_var = tk.StringVar(value=DIRECTION_CONFIG["ka-ru"]["label"])

        self.progress_var = tk.DoubleVar(value=0.0)
        self.status_var = tk.StringVar(value="Готов к работе.")

        self.post_edit_var = tk.BooleanVar(value=False)

        self.start_button: Optional[ttk.Button] = None
        self.post_edit_check: Optional[ttk.Checkbutton] = None

        self.build_ui()

    def build_ui(self):
        pad = 5

        frm = ttk.Frame(self.root, padding=10)
        frm.grid(row=0, column=0, sticky="nsew")

        ttk.Label(frm, text="Файл DOCX/XLSX:").grid(row=0, column=0, sticky="w", pady=pad)
        entry_file = ttk.Entry(frm, textvariable=self.file_path_var, width=60)
        entry_file.grid(row=0, column=1, sticky="we", pady=pad)
        ttk.Button(frm, text="Выбрать...", command=self.choose_file).grid(row=0, column=2, padx=pad, pady=pad)

        ttk.Label(frm, text="Переводчик:").grid(row=1, column=0, sticky="w", pady=pad)

        r1 = ttk.Radiobutton(
            frm,
            text="ChatGPT (облачный)",
            variable=self.translator_var,
            value="chatgpt",
            command=self.on_translator_change,
        )
        r1.grid(row=1, column=1, sticky="w", pady=pad)

        r2 = ttk.Radiobutton(
            frm,
            text="Локальная модель (NLLB-200)",
            variable=self.translator_var,
            value="local",
            command=self.on_translator_change,
        )
        r2.grid(row=2, column=1, sticky="w", pady=pad)

        ttk.Label(frm, text="Направление перевода:").grid(row=3, column=0, sticky="w", pady=pad)
        direction_values = [meta["label"] for meta in DIRECTION_CONFIG.values()]
        self.direction_combo = ttk.Combobox(
            frm,
            textvariable=self.direction_label_var,
            values=direction_values,
            state="readonly",
            width=35,
        )
        self.direction_combo.grid(row=3, column=1, sticky="w", pady=pad)

        ttk.Label(frm, text="Модель ChatGPT:").grid(row=4, column=0, sticky="w", pady=pad)
        self.model_combo = ttk.Combobox(
            frm,
            textvariable=self.model_var,
            values=CHATGPT_MODELS,
            state="readonly",
            width=35,
        )
        self.model_combo.grid(row=4, column=1, sticky="w", pady=pad)

        ttk.Label(frm, text=".env с токеном:").grid(row=5, column=0, sticky="w", pady=pad)
        self.env_entry = ttk.Entry(frm, textvariable=self.env_path_var, width=60)
        self.env_entry.grid(row=5, column=1, sticky="we", pady=pad)
        self.env_button = ttk.Button(frm, text="Выбрать .env...", command=self.choose_env_file)
        self.env_button.grid(row=5, column=2, padx=pad, pady=pad)

        self.post_edit_check = ttk.Checkbutton(
            frm,
            text="Литературная вычитка (улучшать стиль перевода)",
            variable=self.post_edit_var,
        )
        self.post_edit_check.grid(row=6, column=0, columnspan=3, sticky="w", pady=pad)

        ttk.Label(frm, text="Прогресс:").grid(row=7, column=0, sticky="w", pady=pad)
        self.progress_bar = ttk.Progressbar(
            frm,
            maximum=100.0,
            variable=self.progress_var,
            mode="determinate",
            length=300,
        )
        self.progress_bar.grid(row=7, column=1, columnspan=2, sticky="we", pady=pad)

        self.status_label = ttk.Label(frm, textvariable=self.status_var)
        self.status_label.grid(row=8, column=0, columnspan=3, sticky="w", pady=pad)

        self.start_button = ttk.Button(frm, text="Старт перевода", command=self.run_translation)
        self.start_button.grid(row=9, column=0, columnspan=3, pady=10)

        self.root.columnconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)

        self.on_translator_change()

    def _update_progress_mainthread(self, pct: float, msg: str) -> None:
        self.progress_var.set(max(0.0, min(100.0, pct)))
        self.status_var.set(msg)

    def set_progress(self, pct: float, msg: str) -> None:
        self.root.after(0, self._update_progress_mainthread, pct, msg)

    def choose_file(self):
        path = filedialog.askopenfilename(
            title="Выберите DOCX или XLSX",
            filetypes=[("Office files", "*.docx *.xlsx"), ("Все файлы", "*.*")],
        )
        if path:
            self.file_path_var.set(path)

    def choose_env_file(self):
        path = filedialog.askopenfilename(
            title="Выберите .env файл с OPENAI_API_KEY",
            filetypes=[("ENV files", "*.env;*.txt;*.*"), ("Все файлы", "*.*")],
        )
        if path:
            self.env_path_var.set(path)

    def on_translator_change(self):
        kind = self.translator_var.get()
        if kind == "chatgpt":
            self.model_combo.configure(state="readonly")
            self.env_entry.configure(state="normal")
            self.env_button.configure(state="normal")
            if self.post_edit_check is not None:
                self.post_edit_check.configure(
                    text="Литературная вычитка (через ChatGPT)",
                    state="normal",
                )
        else:
            self.model_combo.configure(state="disabled")
            self.env_entry.configure(state="disabled")
            self.env_button.configure(state="disabled")
            if self.post_edit_check is not None:
                self.post_edit_check.configure(
                    text="Литературная вычитка (локальная Qwen2.5-3B-Instruct)",
                    state="normal",
                )

    def run_translation(self):
        file_path = self.file_path_var.get().strip()
        if not file_path:
            messagebox.showerror("Ошибка", "Выберите файл DOCX/XLSX.")
            return

        if not (is_docx(file_path) or is_xlsx(file_path)):
            messagebox.showerror("Ошибка", "Поддерживаются только файлы .docx и .xlsx.")
            return

        translator_kind = self.translator_var.get()
        chatgpt_model = self.model_var.get()
        env_path = self.env_path_var.get().strip() if translator_kind == "chatgpt" else None
        post_edit = bool(self.post_edit_var.get())

        direction_label = self.direction_label_var.get()
        try:
            direction_code = get_direction_code_from_label(direction_label)
        except ValueError as e:
            messagebox.showerror("Ошибка", str(e))
            return

        if translator_kind == "chatgpt" and not env_path:
            messagebox.showerror("Ошибка", "Выберите .env файл с токеном для ChatGPT.")
            return

        self.start_button.configure(state="disabled")
        self.set_progress(0.0, "Начало обработки...")

        t = threading.Thread(
            target=self._worker_translate,
            args=(file_path, translator_kind, chatgpt_model, env_path, direction_code, post_edit),
            daemon=True,
        )
        t.start()

    def _worker_translate(self, file_path: str, translator_kind: str,
                          chatgpt_model: str, env_path: Optional[str],
                          direction_code: str, post_edit: bool):
        try:
            output_path = process_file(
                file_path=file_path,
                translator_kind=translator_kind,
                chatgpt_model=chatgpt_model,
                env_path=env_path,
                direction_code=direction_code,
                post_edit=post_edit,
                progress_callback=self.set_progress,
            )
        except Exception as e:
            import traceback
            traceback.print_exc()

            err_msg = f"{type(e).__name__}: {e}"

            def show_error(msg=err_msg):
                self.start_button.configure(state="normal")
                self._update_progress_mainthread(0.0, "Ошибка.")
                messagebox.showerror("Ошибка", f"Перевод завершился с ошибкой:\n{msg}")

            self.root.after(0, show_error)
            return

        def on_done():
            self.start_button.configure(state="normal")
            self._update_progress_mainthread(100.0, "Готово.")
            messagebox.showinfo("Готово", f"Файл переведён:\n{output_path}")
        self.root.after(0, on_done)


def main():
    root = tk.Tk()
    TranslatorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
