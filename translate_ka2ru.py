import os
import re
import json
import zipfile
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from typing import Dict, List, Set, Iterable, Callable, Optional

import xml.etree.ElementTree as ET

# ============ Настройки по умолчанию ============

DEFAULT_CHATGPT_MODEL = "gpt-4.1-mini"
CHATGPT_MODELS = [
    "gpt-4.1-mini",
    "gpt-4.1",
    "gpt-4o-mini",
    "gpt-4o",
    "gpt-5.1",
    "gpt-5-mini",
]

# направления перевода: грузинский -> целевой язык
DIRECTION_CONFIG: Dict[str, Dict[str, str]] = {
    "ka-ru": {
        "label": "Грузинский → Русский",
        "target_language": "Russian",
        "suffix": "_ru",
        "local_model": "Helsinki-NLP/opus-mt-ka-ru",
    },
    "ka-en": {
        "label": "Грузинский → Английский",
        "target_language": "English",
        "suffix": "_en",
        "local_model": "Helsinki-NLP/opus-mt-ka-en",
    },
}

# диапазоны Unicode для грузинского
GEORGIAN_RE = re.compile(r"[\u10A0-\u10FF\u1C90-\u1CBF]+")


# ============ Утилиты ============

def is_docx(path: str) -> bool:
    return path.lower().endswith(".docx")


def is_xlsx(path: str) -> bool:
    return path.lower().endswith(".xlsx")


def chunks(lst: List[str], n: int) -> Iterable[List[str]]:
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def load_api_key_from_env_file(env_path: str) -> str:
    """
    Читает .env-файл и ищет OPENAI_API_KEY / API_KEY.
    Если не нашёл — берёт первую непустую, некомментированную строку как ключ.
    """
    if not os.path.exists(env_path):
        raise FileNotFoundError(f".env файл не найден: {env_path}")

    candidate = None
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip()
                if k in ("OPENAI_API_KEY", "API_KEY"):
                    return v
                if candidate is None and v:
                    candidate = v
            else:
                if candidate is None:
                    candidate = line

    if not candidate:
        raise ValueError("Не удалось извлечь ключ API из .env файла.")
    return candidate


def get_direction_code_from_label(label: str) -> str:
    for code, meta in DIRECTION_CONFIG.items():
        if meta["label"] == label:
            return code
    raise ValueError(f"Неизвестное направление перевода: {label}")


# ============ Работа с XML внутри DOCX/XLSX ============

def collect_docx_items(path: str) -> List[Dict[str, object]]:
    """
    DOCX — работаем по АБЗАЦАМ (<w:p>), но теперь собираем не просто тексты,
    а ПОЛНЫЙ список элементов с позиционными ID.

    Возвращает список словарей:
      {
        "id": "word/document.xml::p17",
        "xml_name": "word/document.xml",
        "p_index": 17,
        "full_text": "<весь текст абзаца как есть>",
        "clean_text": "<full_text.strip()>"
      }

    Собираем только те абзацы, в которых есть грузинский текст.
    """
    items: List[Dict[str, object]] = []

    with zipfile.ZipFile(path, "r") as zin:
        for info in zin.infolist():
            fname = info.filename
            if not (fname.startswith("word/") and fname.lower().endswith(".xml")):
                continue

            xml_bytes = zin.read(fname)
            try:
                root = ET.fromstring(xml_bytes)
            except Exception:
                continue

            m = re.match(r"\{(.*)\}", root.tag)
            ns = m.group(1) if m else ""
            p_tag = f"{{{ns}}}p"
            t_tag = f"{{{ns}}}t"

            for p_index, p in enumerate(root.iter(p_tag)):
                t_elems = list(p.iter(t_tag))
                if not t_elems:
                    continue

                parts = [(t.text or "") for t in t_elems]
                full_text = "".join(parts)
                if not full_text:
                    continue

                if not GEORGIAN_RE.search(full_text):
                    continue

                clean_text = full_text.strip()
                if not clean_text:
                    continue

                item_id = f"{fname}::p{p_index}"
                items.append({
                    "id": item_id,
                    "xml_name": fname,
                    "p_index": p_index,
                    "full_text": full_text,
                    "clean_text": clean_text,
                })

    print(f"📄 DOCX: найдено {len(items)} абзацев с грузинским текстом.")
    return items


def collect_georgian_fragments_from_xml_bytes(xml_bytes: bytes) -> Set[str]:
    """
    Универсальная функция: находим все текстовые узлы, содержащие грузинский,
    и берём их полный текст целиком (strip()).
    """
    result: Set[str] = set()
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return result

    for elem in root.iter():
        text = elem.text
        if not text:
            continue
        if GEORGIAN_RE.search(text):
            cleaned = text.strip()
            if cleaned:
                result.add(cleaned)
    return result


def collect_fragments_xlsx(path: str) -> Set[str]:
    """
    XLSX — обрабатываем sharedStrings + worksheets.
    workbook.xml (имена листов) не трогаем.
    """
    to_translate: Set[str] = set()

    with zipfile.ZipFile(path, "r") as zin:
        for info in zin.infolist():
            fname = info.filename.lower()

            if fname.startswith("xl/sharedstrings") and fname.endswith(".xml"):
                xml_bytes = zin.read(info.filename)
            elif fname.startswith("xl/worksheets/") and fname.endswith(".xml"):
                xml_bytes = zin.read(info.filename)
            else:
                continue

            frags = collect_georgian_fragments_from_xml_bytes(xml_bytes)
            to_translate.update(frags)

    print(f"📊 XLSX: найдено {len(to_translate)} уникальных грузинских фрагментов.")
    return to_translate


def replace_georgian_in_xml_bytes(xml_bytes: bytes, mapping: Dict[str, str]) -> bytes:
    """
    Для XLSX: если текст узла (strip) есть в mapping — заменяем целиком,
    сохраняя ведущие/хвостовые пробелы.
    """
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return xml_bytes

    for elem in root.iter():
        text = elem.text
        if not text:
            continue

        stripped = text.strip()
        if stripped in mapping:
            prefix_len = len(text) - len(text.lstrip())
            suffix_len = len(text) - len(text.rstrip())
            prefix = text[:prefix_len]
            suffix = text[len(text) - suffix_len:] if suffix_len > 0 else ""
            new_text = mapping[stripped]
            elem.text = f"{prefix}{new_text}{suffix}"

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def process_docx_xml_paragraphs(
    xml_bytes: bytes,
    xml_name: str,
    id_mapping: Dict[str, str],
) -> bytes:
    """
    Пробегаем по <w:p> в одном XML-файле DOCX и, если для параграфа есть
    перевод в id_mapping, подменяем его текст. При этом:
      - если в текущем абзаце уже НЕТ грузинских букв, мы его НЕ трогаем,
        даже если его ID есть в id_mapping (защита от повторного прогона
        по уже переведённому файлу / вручную переведённым кускам).
    """
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return xml_bytes

    m = re.match(r"\{(.*)\}", root.tag)
    ns = m.group(1) if m else ""
    p_tag = f"{{{ns}}}p"
    t_tag = f"{{{ns}}}t"

    for p_index, p in enumerate(root.iter(p_tag)):
        para_id = f"{xml_name}::p{p_index}"
        if para_id not in id_mapping:
            continue

        t_elems = list(p.iter(t_tag))
        if not t_elems:
            continue

        orig_parts = [(t.text or "") for t in t_elems]
        orig_full = "".join(orig_parts)
        if not orig_full:
            continue

        # 🔒 КРИТИЧЕСКОЕ МЕСТО:
        # если В ЭТОМ АБЗАЦЕ уже нет грузинских букв — оставляем как есть
        if not GEORGIAN_RE.search(orig_full):
            continue

        translated_clean = id_mapping[para_id]

        # дебаг
        #print("\n=== RAW TRANSLATION BEFORE SPLIT ===")
        #print("ID:", para_id)
        #print("ORIGINAL:", repr(orig_full))
        #print("TRANSLATED:", repr(translated_clean))
        #print("====================================\n")

        # сохраняем ведущие/хвостовые пробелы абзаца
        lead = len(orig_full) - len(orig_full.lstrip())
        trail = len(orig_full) - len(orig_full.rstrip())
        prefix = orig_full[:lead]
        suffix = orig_full[len(orig_full) - trail:] if trail > 0 else ""
        translated_full = prefix + translated_clean + suffix

        # --- КРИТИЧЕСКОЕ УПРОЩЕНИЕ ---
        # Весь текст абзаца кладём в первый <w:t>, остальные чистим.
        t_elems[0].text = translated_full
        for t in t_elems[1:]:
            t.text = ""
        # ------------------------------

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)

def debug_scan_docx_for_georgian(path: str, max_examples: int = 20) -> None:
    """
    Отладочная функция: сканирует DOCX и ищет все абзацы, где остался грузинский текст.
    Печатает количество и несколько примеров (кусочек текста + имя XML-файла + индекс абзаца).
    """
    count = 0
    examples = []

    with zipfile.ZipFile(path, "r") as zin:
        for info in zin.infolist():
            fname = info.filename
            if not (fname.startswith("word/") and fname.lower().endswith(".xml")):
                continue

            xml_bytes = zin.read(fname)
            try:
                root = ET.fromstring(xml_bytes)
            except Exception:
                continue

            m = re.match(r"\{(.*)\}", root.tag)
            ns = m.group(1) if m else ""
            p_tag = f"{{{ns}}}p"
            t_tag = f"{{{ns}}}t"

            for p_index, p in enumerate(root.iter(p_tag)):
                parts = []
                for t in p.iter(t_tag):
                    parts.append(t.text or "")
                full_text = "".join(parts)
                if not full_text:
                    continue
                if GEORGIAN_RE.search(full_text):
                    count += 1
                    if len(examples) < max_examples:
                        snippet = full_text.strip()
                        if len(snippet) > 120:
                            snippet = snippet[:117] + "..."
                        examples.append((fname, p_index, snippet))

    print(f"🔍 В файле {os.path.basename(path)} осталось абзацев с грузинским: {count}")
    for fname, p_index, snippet in examples:
        print(f"  - {fname}::p{p_index}: {snippet}")


def apply_translations_docx(
    input_path: str,
    output_path: str,
    id_mapping: Dict[str, str],
    text_mapping: Optional[Dict[str, str]] = None,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 90.0,
    end: float = 100.0,
) -> None:
    with zipfile.ZipFile(input_path, "r") as zin, \
         zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:

        infos = zin.infolist()
        total = len(infos)
        changed = 0

        for idx, info in enumerate(infos, start=1):
            fname = info.filename
            data = zin.read(fname)

            new_data = data

            # 1) для word/*.xml прогоняем абзацы с id_mapping
            if fname.startswith("word/") and fname.lower().endswith(".xml"):
                new_data = process_docx_xml_paragraphs(new_data, fname, id_mapping)

            # 2) для ЛЮБОГО *.xml (включая word/*.xml, docProps и т.п.) —
            #    общая замена по text_mapping
            if fname.lower().endswith(".xml") and text_mapping:
                new_data = replace_georgian_in_xml_bytes(new_data, text_mapping)

            if new_data != data:
                changed += 1

            zout.writestr(info, new_data)

            if progress_callback and total > 0:
                frac = idx / total
                pct = start + (end - start) * frac
                progress_callback(pct, "Применение перевода в DOCX...")

    print(f"💾 DOCX сохранён: {output_path}, изменённых XML: {changed}")

def apply_translations_xlsx(
    input_path: str,
    output_path: str,
    text_mapping: Dict[str, str],
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 90.0,
    end: float = 100.0,
) -> None:
    """
    Безопасно применяет переводы к XLSX через openpyxl.

    text_mapping: {исходный_грузинский_текст_strip -> перевод}

    Логика:
      - открываем книгу через openpyxl;
      - для всех ячеек со строковым значением:
          * берём value;
          * делаем stripped = value.strip();
          * если stripped есть в text_mapping — подменяем, аккуратно сохраняя
            ведущие/хвостовые пробелы;
      - сохраняем в output_path.
    """
    from openpyxl import load_workbook

    # Открываем книгу
    wb = load_workbook(input_path, data_only=False)

    # Считаем общее число ячеек для более-менее честного прогресса
    total_cells = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            total_cells += len(row)
    if total_cells == 0:
        total_cells = 1  # защита от деления на ноль

    processed = 0
    changed_cells = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str):
                    original = val
                    stripped = val.strip()
                    if stripped in text_mapping:
                        new_core = text_mapping[stripped]

                        # сохраняем ведущие/хвостовые пробелы
                        prefix_len = len(original) - len(original.lstrip())
                        suffix_len = len(original) - len(original.rstrip())
                        prefix = original[:prefix_len]
                        suffix = original[len(original) - suffix_len:] if suffix_len > 0 else ""

                        cell.value = f"{prefix}{new_core}{suffix}"
                        changed_cells += 1

                processed += 1
                if progress_callback:
                    frac = processed / total_cells
                    pct = start + (end - start) * frac
                    progress_callback(pct, "Применение перевода в XLSX...")

    wb.save(output_path)
    print(f"💾 XLSX сохранён: {output_path}, изменённых ячеек: {changed_cells}")



# ============ Переводчики ============

def translate_with_chatgpt(
    fragments: List[str],
    model_name: str,
    api_key: str,
    target_language: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 10.0,
    end: float = 90.0,
) -> Dict[str, str]:
    """
    Перевод грузинских фрагментов через ChatGPT (OpenAI API).
    НОВАЯ ВЕРСИЯ:
      - не использует исходные тексты как JSON-ключи;
      - вместо этого шлёт список объектов {id, text};
      - модель возвращает translations: [{id, text}, ...];
      - на основе этого строим mapping {оригинал: перевод}.
    """
    from openai import OpenAI

    os.environ["OPENAI_API_KEY"] = api_key
    client = OpenAI()

    # 1) Чистим и дедуплируем тексты
    cleaned = []
    for f in fragments:
        s = (f or "").strip()
        if s:
            cleaned.append(s)

    # уникальные тексты, чтобы не платить дважды за одинаковые абзацы
    unique_texts: List[str] = []
    seen: Set[str] = set()
    for s in cleaned:
        if s not in seen:
            seen.add(s)
            unique_texts.append(s)

    if not unique_texts:
        return {}

    # Присваиваем каждому уникальному тексту числовой ID
    id_to_text: Dict[int, str] = {i: txt for i, txt in enumerate(unique_texts)}
    text_to_id: Dict[str, int] = {txt: i for i, txt in id_to_text.items()}

    print(f"Для перевода через ChatGPT подготовлено {len(unique_texts)} уникальных фрагментов.")

    BATCH_SIZE = 200
    total = len(unique_texts)
    done = 0

    # сюда будем собирать переводы по ID
    id_to_translated: Dict[int, str] = {}

    for batch_ids in chunks(list(id_to_text.keys()), BATCH_SIZE):
        batch_items = [
            {"id": i, "text": id_to_text[i]}
            for i in batch_ids
        ]

        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Перевод через ChatGPT...")

        user_payload = {
            "source_language": "Georgian",
            "target_language": target_language,
            "items": batch_items,
        }

        system_msg = (
            "You are a professional legal and technical translator. "
            "The texts are official documents (tariff methodology, regulatory acts, explanatory notes). "
            f"Translate from Georgian to {target_language} into natural, formal, human-quality {target_language}. "
            "You MAY freely change word order, grammar, and morphology so that the result sounds like good native legal language, "
            "but you MUST preserve all facts, numbers, names, and logical relations. "
            "Avoid literal calques from Georgian where they sound unnatural in the target language. "
            "DO NOT merge separate words together: always keep proper spaces between words, "
            "between prepositions and nouns, and around conjunctions (like 'и', 'და', 'and', etc.). "
            "Fix any missing spaces if they are present in the original. "
            "Do NOT explicitly mention grammatical cases or parts of speech. "
            "Return ONLY a JSON object with a single key 'translations', whose value is a list of objects "
            "of the form {\"id\": <same id>, \"text\": <translation>}. "
            "Do not add extra fields."
        )

        resp = client.chat.completions.create(
            model=model_name,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
            ],
        )

        content = resp.choices[0].message.content
        try:
            data = json.loads(content)
        except json.JSONDecodeError:
            print("⚠️ Ошибка JSON от модели, ответ:")
            print(content)
            raise

        translations_list = data.get("translations")
        if not isinstance(translations_list, list):
            raise ValueError("Ожидался ключ 'translations' со списком объектов {id, text}.")

        for obj in translations_list:
            if not isinstance(obj, dict):
                continue
            tid = obj.get("id")
            ttext = obj.get("text")
            try:
                tid_int = int(tid)
            except (TypeError, ValueError):
                continue
            if not isinstance(ttext, str) or not ttext.strip():
                # если модель не дала перевод — оставим как есть (подставим позже)
                continue
            id_to_translated[tid_int] = ttext

        done += len(batch_ids)
        print(f"   ChatGPT перевёл {done}/{total} уникальных фрагментов")

        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Перевод через ChatGPT...")

    # 2) Собираем окончательный mapping {исходный_текст -> перевод}
    mapping: Dict[str, str] = {}
    for txt, tid in text_to_id.items():
        trans = id_to_translated.get(tid)
        if isinstance(trans, str) and trans.strip():
            # НИЧЕГО не правим: используем перевод как есть
            mapping[txt] = trans
        else:
            mapping[txt] = txt

    return mapping


def post_edit_with_chatgpt(
    mapping: Dict[str, str],
    model_name: str,
    api_key: str,
    target_language: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 70.0,
    end: float = 90.0,
) -> Dict[str, str]:
    """
    Литературная вычитка уже переведённого текста.
    mapping: {грузинский_оригинал -> машинный_перевод}
    Возвращает тот же mapping, но значения сглажены.
    """
    from openai import OpenAI

    os.environ["OPENAI_API_KEY"] = api_key
    client = OpenAI()

    unique_values: List[str] = []
    seen = set()
    for v in mapping.values():
        if v not in seen and v.strip():
            seen.add(v)
            unique_values.append(v)

    if not unique_values:
        return mapping

    BATCH_SIZE = 200
    total = len(unique_values)
    done = 0

    improved_map: Dict[str, str] = {}

    for batch in chunks(unique_values, BATCH_SIZE):
        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Литературная вычитка перевода...")

        user_payload = {
            "target_language": target_language,
            "texts": batch,
        }

        system_msg = (
            "You are a professional editor for legal, regulatory and technical documents. "
            f"Improve style, clarity, grammar and fluency in {target_language} while preserving the same meaning, facts, numbers and legal content. "
            "You MAY change word order, fix awkward literal phrases, adjust cases and morphology, "
            "replace unnatural calques with standard legal expressions, and break or merge sentences if it improves readability. "
            "In addition, you MUST carefully fix spacing: "
            "add missing spaces between words, between prepositions and the following words, "
            "between numbers and words (like 'от 4 декабря', '№ 33'), and after punctuation marks where appropriate. "
            "Do NOT merge distinct words together. "
            "Do NOT add new facts or remove existing ones. "
            "Avoid explicit linguistic labels like 'родительный падеж' or explanations of grammar. "
            "Return ONLY a JSON object mapping each original text to its improved version. "
            "Keys MUST be EXACTLY the original texts. Do not add extra fields."
        )

        resp = client.chat.completions.create(
            model=model_name,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
            ],
        )

        content = resp.choices[0].message.content
        try:
            data = json.loads(content)
        except json.JSONDecodeError:
            print("⚠️ Ошибка JSON от модели (вычитка), ответ:")
            print(content)
            raise

        for orig_text in batch:
            new_text = data.get(orig_text)
            if isinstance(new_text, str) and new_text.strip():
                # НЕ трогаем пробелы регэксами, берём как есть
                improved_map[orig_text] = new_text
            else:
                improved_map[orig_text] = orig_text

        done += len(batch)
        print(f"   ChatGPT вычитал {done}/{total} фрагментов")

        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Литературная вычитка перевода...")

    new_mapping: Dict[str, str] = {}
    for geo, raw_trans in mapping.items():
        new_mapping[geo] = improved_map.get(raw_trans, raw_trans)

    return new_mapping

def fix_spacing_with_chatgpt(
    mapping: Dict[str, str],
    model_name: str,
    api_key: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 70.0,
    end: float = 90.0,
) -> Dict[str, str]:
    """
    Аккуратная правка ПРОБЕЛОВ в уже переведённом русском тексте.
    Важно: модель НЕ ИМЕЕТ ПРАВА менять какие-либо символы, кроме обычных пробелов U+0020.
    mapping: {грузинский_оригинал -> русский_перевод}
    Возвращает mapping с теми же ключами, но значения с поправленными пробелами.
    """
    from openai import OpenAI

    os.environ["OPENAI_API_KEY"] = api_key
    client = OpenAI()

    # Собираем уникальные русские строки
    unique_values: List[str] = []
    seen = set()
    for v in mapping.values():
        if isinstance(v, str) and v not in seen and v.strip():
            seen.add(v)
            unique_values.append(v)

    if not unique_values:
        return mapping

    BATCH_SIZE = 200
    total = len(unique_values)
    done = 0

    fixed_map: Dict[str, str] = {}

    for batch in chunks(unique_values, BATCH_SIZE):
        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Правка пробелов в русском тексте...")

        user_payload = {
            "texts": batch,
        }

        system_msg = (
            "You receive Russian texts which are already translated correctly. "
            "Your ONLY task is to fix spacing errors: insert or delete ASCII space characters (U+0020) "
            "where necessary between words, numbers and punctuation, and collapse multiple spaces to single ones if appropriate. "
            "You MUST NOT change, delete, reorder or insert ANY non-space characters (letters, digits, punctuation). "
            "The sequence of all non-space characters must remain EXACTLY the same and in the same order. "
            "Return ONLY a JSON object mapping each original string to its corrected version. "
            "Keys MUST be EXACTLY the original strings. Do not add extra fields."
        )

        resp = client.chat.completions.create(
            model=model_name,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
            ],
        )

        content = resp.choices[0].message.content
        try:
            data = json.loads(content)
        except json.JSONDecodeError:
            print("⚠️ Ошибка JSON от модели (fix_spacing), ответ:")
            print(content)
            raise

        # data: {original_text -> fixed_text}
        for orig_text in batch:
            new_text = data.get(orig_text)
            if isinstance(new_text, str) and new_text.strip():
                fixed_map[orig_text] = new_text
            else:
                fixed_map[orig_text] = orig_text

        done += len(batch)
        print(f"   ChatGPT поправил пробелы в {done}/{total} фрагментах")

        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Правка пробелов в русском тексте...")

    # Собираем новый mapping: грузинский -> русский(с нормальными пробелами)
    new_mapping: Dict[str, str] = {}
    for geo, ru in mapping.items():
        if isinstance(ru, str):
            new_mapping[geo] = fixed_map.get(ru, ru)
        else:
            new_mapping[geo] = ru

    return new_mapping


def translate_with_local_model(
    fragments: List[str],
    direction_code: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 10.0,
    end: float = 90.0,
) -> Dict[str, str]:
    """
    Локальный перевод через Helsinki-NLP в зависимости от направления:
    ka-ru -> opus-mt-ka-ru
    ka-en -> opus-mt-ka-en
    Требуется: pip install transformers torch sentencepiece
    """
    from transformers import MarianMTModel, MarianTokenizer
    import torch

    meta = DIRECTION_CONFIG[direction_code]
    MODEL_NAME = meta["local_model"]

    print(f"⏳ Загружаем локальную модель ({MODEL_NAME})...")
    if progress_callback:
        progress_callback(start, "Загрузка локальной модели...")

    tokenizer = MarianTokenizer.from_pretrained(MODEL_NAME)
    model = MarianMTModel.from_pretrained(MODEL_NAME)

    mapping: Dict[str, str] = {}
    BATCH_SIZE = 64

    remaining = [f for f in fragments if f.strip()]
    total = len(remaining)
    done = 0

    for batch in chunks(remaining, BATCH_SIZE):
        inputs = tokenizer(batch, return_tensors="pt", padding=True, truncation=True)
        with torch.no_grad():
            generated = model.generate(**inputs, max_length=512)
        outputs = tokenizer.batch_decode(generated, skip_special_tokens=True)

        for orig, trans in zip(batch, outputs):
            mapping[orig] = trans if trans.strip() else orig

        done += len(batch)
        print(f"   Локальная модель перевела {done}/{total} фрагментов")

        if progress_callback and total > 0:
            frac = done / total
            pct = start + (end - start) * frac
            progress_callback(pct, "Перевод локальной моделью...")

    return mapping


def normalize_segment_boundaries(segments: List[str]) -> List[str]:
    """
    Очень аккуратная нормализация границ:
    - НЕ добавляет новые пробелы там, где их не было;
    - только:
        * схлопывает пачки пробелов в конце/начале сегментов;
        * если left заканчивается пробелом и right начинается пробелом —
          убирает пробелы в начале right (оставляя один со стороны left).
    """
    if len(segments) <= 1:
        return segments

    segs = segments[:]

    for i in range(len(segs) - 1):
        left = segs[i]
        right = segs[i + 1]

        # схлопываем пачки пробелов ВНУТРИ сегментов и на краях
        left = re.sub(r' {2,}', ' ', left)
        right = re.sub(r' {2,}', ' ', right)

        # если left заканчивается пробелом и right начинается пробелом —
        # оставляем один (со стороны left)
        if left.endswith(" ") and right.startswith(" "):
            right = right.lstrip()

        segs[i] = left
        segs[i + 1] = right

    return segs

def fix_basic_spacing_ru(text: str) -> str:
    """
    Лёгкая правка очевидных случаев типа '№33', '2020года', запятая без пробела и т.п.
    Без словарей, только простые паттерны.
    """
    import re

    text = text.replace("\u00A0", " ")

    # "Грузии№33" -> "Грузии №33"
    text = re.sub(r'([А-ЯЁа-яё])№\s*(\d)', r'\1 №\2', text)
    text = re.sub(r'№\s*(\d)', r'№ \1', text)

    # предлог + число ("от4" -> "от 4")
    text = re.sub(r'(?i)\b(от|до|по|на|в|к|с|у)(\d)', r'\1 \2', text)

    # число + слово ("4декабря" -> "4 декабря")
    text = re.sub(r'(\d)([А-ЯЁа-яё])', r'\1 \2', text)

    # число + "год/годы/гг."
    text = re.sub(r'(\d{3,4})\s*(год[ауе]?|гг?\.?)', r'\1 \2', text)

    # знаки препинания без пробела после
    text = re.sub(r',([^\s])', r', \1', text)
    text = re.sub(r';([^\s])', r'; \1', text)
    text = re.sub(r'([^.])\.([А-ЯЁ])', r'\1. \2', text)

    # проценты
    text = re.sub(r'(%)([А-ЯЁа-яё])', r'\1 \2', text)

    # схлопываем длинные пачки пробелов
    text = re.sub(r'[ \t]{2,}', ' ', text)

    return text


# ============ Логика перевода одного файла ============

def process_file(
    file_path: str,
    translator_kind: str,
    chatgpt_model: str,
    env_path: Optional[str],
    direction_code: str,
    post_edit: bool,
    progress_callback: Optional[Callable[[float, str], None]] = None,
) -> str:
    """
    Главная функция: собирает фрагменты, переводит выбранным движком,
    при необходимости делает литературную вычитку,
    применяет переводы, возвращает путь к выходному файлу.
    """

    if progress_callback is None:
        def progress_callback(pct: float, msg: str) -> None:
            pass  # заглушка

    if not (is_docx(file_path) or is_xlsx(file_path)):
        raise ValueError("Поддерживаются только файлы .docx и .xlsx")

    if direction_code not in DIRECTION_CONFIG:
        raise ValueError(f"Неизвестное направление: {direction_code}")

    meta = DIRECTION_CONFIG[direction_code]
    target_language = meta["target_language"]
    suffix = meta["suffix"]

    # 1. Сбор фрагментов
    progress_callback(0.0, "Сбор грузинских фрагментов...")

    if is_docx(file_path):
        items = collect_docx_items(file_path)
        if not items:
            progress_callback(0.0, "Грузинский текст не найден.")
            raise RuntimeError("В файле не найден грузинский текст для перевода.")

        # 1) абзацы (как и было)
        base_texts = {str(it["clean_text"]) for it in items}

        # 2) ДОПОЛНИТЕЛЬНО: любые грузинские фрагменты из всех *.xml внутри docx
        extra_texts: Set[str] = set()
        with zipfile.ZipFile(file_path, "r") as zin:
            for info in zin.infolist():
                fname = info.filename
                if not fname.lower().endswith(".xml"):
                    continue

                xml_bytes = zin.read(fname)
                extra_texts.update(collect_georgian_fragments_from_xml_bytes(xml_bytes))

        # объединяем
        all_texts = base_texts | extra_texts

        # на всякий случай ещё раз фильтруем по грузинскому (если вдруг что-то пролезло)
        fragments_for_translation = sorted(
            t for t in all_texts
            if t.strip() and GEORGIAN_RE.search(t)
        )

        items_for_docx = items  # сохраним, чтобы потом построить id_mapping
    else:
        # XLSX — как раньше, просто множество строк
        fragments_set = collect_fragments_xlsx(file_path)
        if not fragments_set:
            progress_callback(0.0, "Грузинский текст не найден.")
            raise RuntimeError("В файле не найден грузинский текст для перевода.")
        fragments_for_translation = sorted(fragments_set)
        items = None  # для XLSX не нужно

    print(f"Найдено {len(fragments_for_translation)} уникальных фрагментов для перевода.")
    progress_callback(5.0, f"Найдено {len(fragments_for_translation)} фрагментов. Подготовка к переводу...")

    # 2. Перевод
    if translator_kind == "chatgpt":
        if not env_path:
            raise ValueError("Не выбран .env файл с токеном для ChatGPT.")
        api_key = load_api_key_from_env_file(env_path)

        mapping_text_to_trans = translate_with_chatgpt(
            fragments_for_translation,
            chatgpt_model,
            api_key,
            target_language,
            progress_callback=progress_callback,
            start=10.0,
            end=60.0,
        )

        # Больше НИКАКИХ автоматических правок пробелов здесь.
        # Если хочешь, пост-редактуру можно включать отдельно (но она может менять пробелы).
        if post_edit:
            mapping_text_to_trans = post_edit_with_chatgpt(
                mapping_text_to_trans,
                chatgpt_model,
                api_key,
                target_language,
                progress_callback=progress_callback,
                start=60.0,
                end=90.0,
            )

    else:
        # Локальный перевод без ChatGPT-постобработки
        mapping_text_to_trans = translate_with_local_model(
            fragments_for_translation,
            direction_code,
            progress_callback=progress_callback,
            start=10.0,
            end=90.0,
        )

    # 2b. Преобразуем маппинг для DOCX в формат id -> translation
    if is_docx(file_path):
        id_mapping: Dict[str, str] = {}
        for it in items:  # type: ignore
            clean_text = str(it["clean_text"])
            item_id = str(it["id"])
            translated = mapping_text_to_trans.get(clean_text, clean_text)
            id_mapping[item_id] = translated
    else:
        id_mapping = mapping_text_to_trans  # для XLSX используем текстовый mapping

    # 3. Применяем переводы к файлу
    base, ext = os.path.splitext(file_path)
    output_path = f"{base}{suffix}{ext}"

    progress_callback(90.0, "Применяем переводы к файлу...")
    if is_docx(file_path):
        apply_translations_docx(
            file_path,
            output_path,
            id_mapping,
            mapping_text_to_trans,  # <-- вот это важное
            progress_callback=progress_callback,
            start=90.0,
            end=100.0,
        )
    else:
        apply_translations_xlsx(
            file_path,
            output_path,
            id_mapping,
            progress_callback=progress_callback,
            start=90.0,
            end=100.0,
        )

    progress_callback(100.0, "Готово.")
    if is_docx(output_path):
        debug_scan_docx_for_georgian(output_path)

    return output_path


# ============ GUI (Tkinter) ============

class TranslatorGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Перевод грузинского текста в DOCX/XLSX")

        self.file_path_var = tk.StringVar()
        self.env_path_var = tk.StringVar()
        self.translator_var = tk.StringVar(value="chatgpt")
        self.model_var = tk.StringVar(value=DEFAULT_CHATGPT_MODEL)
        self.direction_label_var = tk.StringVar(value=DIRECTION_CONFIG["ka-ru"]["label"])

        self.progress_var = tk.DoubleVar(value=0.0)
        self.status_var = tk.StringVar(value="Готов к работе.")

        self.post_edit_var = tk.BooleanVar(value=False)  # чекбокс вычитки

        self.start_button: Optional[ttk.Button] = None
        self.post_edit_check: Optional[ttk.Checkbutton] = None

        self.build_ui()

    def build_ui(self):
        pad = 5

        frm = ttk.Frame(self.root, padding=10)
        frm.grid(row=0, column=0, sticky="nsew")

        # --- выбор файла ---
        ttk.Label(frm, text="Файл DOCX/XLSX:").grid(row=0, column=0, sticky="w", pady=pad)
        entry_file = ttk.Entry(frm, textvariable=self.file_path_var, width=60)
        entry_file.grid(row=0, column=1, sticky="we", pady=pad)
        ttk.Button(frm, text="Выбрать...", command=self.choose_file).grid(row=0, column=2, padx=pad, pady=pad)

        # --- выбор типа переводчика ---
        ttk.Label(frm, text="Переводчик:").grid(row=1, column=0, sticky="w", pady=pad)

        r1 = ttk.Radiobutton(
            frm,
            text="ChatGPT (облачный)",
            variable=self.translator_var,
            value="chatgpt",
            command=self.on_translator_change,
        )
        r1.grid(row=1, column=1, sticky="w", pady=pad)

        r2 = ttk.Radiobutton(
            frm,
            text="Локальная модель (Helsinki-NLP)",
            variable=self.translator_var,
            value="local",
            command=self.on_translator_change,
        )
        r2.grid(row=2, column=1, sticky="w", pady=pad)

        # --- направление перевода ---
        ttk.Label(frm, text="Направление перевода:").grid(row=3, column=0, sticky="w", pady=pad)
        direction_values = [meta["label"] for meta in DIRECTION_CONFIG.values()]
        self.direction_combo = ttk.Combobox(
            frm,
            textvariable=self.direction_label_var,
            values=direction_values,
            state="readonly",
            width=35,
        )
        self.direction_combo.grid(row=3, column=1, sticky="w", pady=pad)

        # --- выбор модели ChatGPT ---
        ttk.Label(frm, text="Модель ChatGPT:").grid(row=4, column=0, sticky="w", pady=pad)
        self.model_combo = ttk.Combobox(
            frm,
            textvariable=self.model_var,
            values=CHATGPT_MODELS,
            state="readonly",
            width=35,
        )
        self.model_combo.grid(row=4, column=1, sticky="w", pady=pad)

        # --- выбор .env с токеном ---
        ttk.Label(frm, text=".env с токеном:").grid(row=5, column=0, sticky="w", pady=pad)
        self.env_entry = ttk.Entry(frm, textvariable=self.env_path_var, width=60)
        self.env_entry.grid(row=5, column=1, sticky="we", pady=pad)
        self.env_button = ttk.Button(frm, text="Выбрать .env...", command=self.choose_env_file)
        self.env_button.grid(row=5, column=2, padx=pad, pady=pad)

        # --- чекбокс литературной вычитки ---
        self.post_edit_check = ttk.Checkbutton(
            frm,
            text="Литературная вычитка (улучшать стиль перевода через ChatGPT)",
            variable=self.post_edit_var,
        )
        self.post_edit_check.grid(row=6, column=0, columnspan=3, sticky="w", pady=pad)

        # --- прогресс и статус ---
        ttk.Label(frm, text="Прогресс:").grid(row=7, column=0, sticky="w", pady=pad)
        self.progress_bar = ttk.Progressbar(
            frm,
            maximum=100.0,
            variable=self.progress_var,
            mode="determinate",
            length=300,
        )
        self.progress_bar.grid(row=7, column=1, columnspan=2, sticky="we", pady=pad)

        self.status_label = ttk.Label(frm, textvariable=self.status_var)
        self.status_label.grid(row=8, column=0, columnspan=3, sticky="w", pady=pad)

        # --- кнопка запуска ---
        self.start_button = ttk.Button(frm, text="Старт перевода", command=self.run_translation)
        self.start_button.grid(row=9, column=0, columnspan=3, pady=10)

        self.root.columnconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)

        self.on_translator_change()

    # обновление GUI — ТОЛЬКО из главного потока
    def _update_progress_mainthread(self, pct: float, msg: str) -> None:
        self.progress_var.set(max(0.0, min(100.0, pct)))
        self.status_var.set(msg)

    def set_progress(self, pct: float, msg: str) -> None:
        # вызывается из рабочего потока → прокидываем через after
        self.root.after(0, self._update_progress_mainthread, pct, msg)

    def choose_file(self):
        path = filedialog.askopenfilename(
            title="Выберите DOCX или XLSX",
            filetypes=[("Office files", "*.docx *.xlsx"), ("Все файлы", "*.*")],
        )
        if path:
            self.file_path_var.set(path)

    def choose_env_file(self):
        path = filedialog.askopenfilename(
            title="Выберите .env файл с OPENAI_API_KEY",
            filetypes=[("ENV files", "*.env;*.txt;*.*"), ("Все файлы", "*.*")],
        )
        if path:
            self.env_path_var.set(path)

    def on_translator_change(self):
        kind = self.translator_var.get()
        if kind == "chatgpt":
            self.model_combo.configure(state="readonly")
            self.env_entry.configure(state="normal")
            self.env_button.configure(state="normal")
            if self.post_edit_check is not None:
                self.post_edit_check.configure(state="normal")
        else:
            self.model_combo.configure(state="disabled")
            self.env_entry.configure(state="disabled")
            self.env_button.configure(state="disabled")
            self.post_edit_var.set(False)
            if self.post_edit_check is not None:
                self.post_edit_check.configure(state="disabled")

    def run_translation(self):
        file_path = self.file_path_var.get().strip()
        if not file_path:
            messagebox.showerror("Ошибка", "Выберите файл DOCX/XLSX.")
            return

        if not (is_docx(file_path) or is_xlsx(file_path)):
            messagebox.showerror("Ошибка", "Поддерживаются только файлы .docx и .xlsx.")
            return

        translator_kind = self.translator_var.get()
        chatgpt_model = self.model_var.get()
        env_path = self.env_path_var.get().strip() if translator_kind == "chatgpt" else None
        post_edit = bool(self.post_edit_var.get())

        direction_label = self.direction_label_var.get()
        try:
            direction_code = get_direction_code_from_label(direction_label)
        except ValueError as e:
            messagebox.showerror("Ошибка", str(e))
            return

        if translator_kind == "chatgpt" and not env_path:
            messagebox.showerror("Ошибка", "Выберите .env файл с токеном для ChatGPT.")
            return

        # блокируем кнопку и сбрасываем прогресс
        self.start_button.configure(state="disabled")
        self.set_progress(0.0, "Начало обработки...")

        # запускаем перевод в отдельном потоке
        t = threading.Thread(
            target=self._worker_translate,
            args=(file_path, translator_kind, chatgpt_model, env_path, direction_code, post_edit),
            daemon=True,
        )
        t.start()

    def _worker_translate(self, file_path: str, translator_kind: str,
                          chatgpt_model: str, env_path: Optional[str],
                          direction_code: str, post_edit: bool):
        """Рабочий поток: запускает process_file и шлёт прогресс в GUI."""
        try:
            output_path = process_file(
                file_path=file_path,
                translator_kind=translator_kind,
                chatgpt_model=chatgpt_model,
                env_path=env_path,
                direction_code=direction_code,
                post_edit=post_edit,
                progress_callback=self.set_progress,
            )
        except Exception as e:
            # Сохраняем текст ошибки, потому что переменную e потом удалят
            import traceback
            traceback.print_exc()  # чтобы видеть полный трейс в консоли

            err_msg = f"{type(e).__name__}: {e}"

            def show_error(msg=err_msg):
                self.start_button.configure(state="normal")
                self._update_progress_mainthread(0.0, "Ошибка.")
                messagebox.showerror("Ошибка", f"Перевод завершился с ошибкой:\n{msg}")

            self.root.after(0, show_error)
            return

        def on_done():
            self.start_button.configure(state="normal")
            self._update_progress_mainthread(100.0, "Готово.")
            messagebox.showinfo("Готово", f"Файл переведён:\n{output_path}")
        self.root.after(0, on_done)


def main():
    root = tk.Tk()
    TranslatorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
