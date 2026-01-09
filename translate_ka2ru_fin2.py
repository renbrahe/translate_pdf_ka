import os
import re
import json
import time
import zipfile
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from typing import Dict, List, Set, Iterable, Callable, Optional, Any, Tuple

import xml.etree.ElementTree as ET

from openai import RateLimitError
from transformers import AutoTokenizer, AutoModelForSeq2SeqLM, AutoModelForCausalLM
import torch


# =============================================================================
# HuggingFace ONLINE/OFFLINE helpers
# =============================================================================

def hf_enable_online() -> None:
    os.environ.pop("HF_HUB_OFFLINE", None)
    os.environ.pop("TRANSFORMERS_OFFLINE", None)


def hf_enable_offline() -> None:
    os.environ["HF_HUB_OFFLINE"] = "1"
    os.environ["TRANSFORMERS_OFFLINE"] = "1"
    os.environ["HF_HUB_DISABLE_TELEMETRY"] = "1"


def hf_print_mode(prefix: str = "") -> None:
    print(
        f"{prefix}HF_HUB_OFFLINE={os.getenv('HF_HUB_OFFLINE')} "
        f"TRANSFORMERS_OFFLINE={os.getenv('TRANSFORMERS_OFFLINE')}"
    )


# =============================================================================
# Token split (для ChatGPT батчей)
# =============================================================================

def estimate_tokens(text: str) -> int:
    if not text:
        return 1
    return int(len(text) * 1.2)


def split_fragments_by_tokens(
    fragments: List[Dict[str, Any]],
    max_tokens_per_batch: int = 8000,
) -> Iterable[List[Dict[str, Any]]]:
    batch: List[Dict[str, Any]] = []
    current_tokens = 0
    for frag in fragments:
        text = frag["text"]
        t = estimate_tokens(text)

        if t > max_tokens_per_batch:
            if batch:
                yield batch
                batch = []
                current_tokens = 0
            yield [frag]
            continue

        if batch and current_tokens + t > max_tokens_per_batch:
            yield batch
            batch = []
            current_tokens = 0

        batch.append(frag)
        current_tokens += t

    if batch:
        yield batch


def split_texts_by_tokens(
    texts: List[str],
    max_tokens_per_batch: int = 8000,
) -> Iterable[List[str]]:
    batch: List[str] = []
    current_tokens = 0
    for text in texts:
        if not isinstance(text, str):
            continue
        t = estimate_tokens(text)

        if t > max_tokens_per_batch:
            if batch:
                yield batch
                batch = []
                current_tokens = 0
            yield [text]
            continue

        if batch and current_tokens + t > max_tokens_per_batch:
            yield batch
            batch = []
            current_tokens = 0

        batch.append(text)
        current_tokens += t

    if batch:
        yield batch


# =============================================================================
# Настройки
# =============================================================================

DEFAULT_CHATGPT_MODEL = "gpt-4.1-mini"
CHATGPT_MODELS = [
    "gpt-4.1-mini",
    "gpt-4.1",
    "gpt-4o-mini",
    "gpt-4o",
    "gpt-5.1",
    "gpt-5-mini",
]

DIRECTION_CONFIG: Dict[str, Dict[str, str]] = {
    "ka-ru": {"label": "Грузинский → Русский", "target_language": "Russian", "suffix": "_ru"},
    "ka-en": {"label": "Грузинский → Английский", "target_language": "English", "suffix": "_en"},
}

GEORGIAN_RE = re.compile(r"[\u10A0-\u10FF\u1C90-\u1CBF]+")

# Канонические имена (не склоняются)
CANON_TELMICO_RU = "ТЭЛМИКО"
CANON_TELASI_RU = "Теласи"
TELMICO_KA_FORMS = ["თელმიკო", "თელმიკო-ს", "თელმიკოის", "თელმიკო-სთვის"]
TELASI_KA_FORMS = ["თელასი", "თელასი-ს", "თელასის", "თელასი-სთვის"]


# =============================================================================
# Утилиты
# =============================================================================

def is_docx(path: str) -> bool:
    return path.lower().endswith(".docx")


def is_xlsx(path: str) -> bool:
    return path.lower().endswith(".xlsx")


def is_pptx(path: str) -> bool:
    return path.lower().endswith(".pptx")


def chunks(lst: List[str], n: int) -> Iterable[List[str]]:
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def load_api_key_from_env_file(env_path: str) -> str:
    if not os.path.exists(env_path):
        raise FileNotFoundError(f".env файл не найден: {env_path}")

    candidate = None
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip()
                if k in ("OPENAI_API_KEY", "API_KEY"):
                    return v
                if candidate is None and v:
                    candidate = v
            else:
                if candidate is None:
                    candidate = line

    if not candidate:
        raise ValueError("Не удалось извлечь ключ API из .env файла.")
    return candidate


def get_direction_code_from_label(label: str) -> str:
    for code, meta in DIRECTION_CONFIG.items():
        if meta["label"] == label:
            return code
    raise ValueError(f"Неизвестное направление перевода: {label}")


# =============================================================================
# Детерминированные фиксы и валидации (правильные этапы)
# =============================================================================

_RE_KA_BLOCK = re.compile(r"[\u10A0-\u10FF]")
_RE_LATIN_RN = re.compile(r"\br/n\b", re.IGNORECASE)

_RE_VAT_BAD_DN = re.compile(r"\bДН\b", re.IGNORECASE)
_RE_VAT_KA = re.compile(r"\bდღგ(?:-ს|-ის)?\b", flags=re.IGNORECASE)

_RE_JUNK_BIBLE = re.compile(r"\b(Деяни[яе]|Бытие|Псалтир|Евангел|Апостол)\b", re.IGNORECASE)
_RE_NONSENSE_PIRG = re.compile(r"\bпергасаточ\w*\b|\bпергазмаплат\w*\b", re.IGNORECASE)
_RE_CASINO = re.compile(r"\bказино\b", re.IGNORECASE)
_RE_DEPT_DEFENSE = re.compile(r"\bДепартамент\s+обороны\b", re.IGNORECASE)

_RE_DIGIT_PAREN = re.compile(r"(\b\d{1,3}\b)\s*\(\s*([^)]+?)\s*\)")


def _has_ka(s: str) -> bool:
    return bool(_RE_KA_BLOCK.search(s or ""))


def _normalize_spaces(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = s.replace("\u00A0", " ")
    s = re.sub(r"[ \t]{2,}", " ", s)
    return s.strip()


def apply_term_normalization(text: str) -> str:
    if not isinstance(text, str) or not text:
        return text or ""
    s = text

    # VAT
    s = _RE_VAT_BAD_DN.sub("НДС", s)
    s = _RE_VAT_KA.sub("НДС", s)

    # r/n -> п/н
    s = _RE_LATIN_RN.sub("п/н", s)

    # TELMICO / Telasi грузинские формы -> канон
    for f in TELMICO_KA_FORMS:
        s = re.sub(re.escape(f), CANON_TELMICO_RU, s, flags=re.IGNORECASE)
    for f in TELASI_KA_FORMS:
        s = re.sub(re.escape(f), CANON_TELASI_RU, s, flags=re.IGNORECASE)

    # nonsense -> нормальный юридический термин
    s = _RE_NONSENSE_PIRG.sub("неустойка", s)

    # casino -> treasury
    s = re.sub(r"(?i)\bединый\s+счет\s+казино\b", "единый казначейский счет", s)
    s = _RE_CASINO.sub("казначейский", s)

    # department defense artifact
    s = _RE_DEPT_DEFENSE.sub("Департамент полиции охраны", s)

    return _normalize_spaces(s)


_RU_NUM_0_20 = {
    0: "ноль", 1: "один", 2: "два", 3: "три", 4: "четыре", 5: "пять",
    6: "шесть", 7: "семь", 8: "восемь", 9: "девять", 10: "десять",
    11: "одиннадцать", 12: "двенадцать", 13: "тринадцать", 14: "четырнадцать",
    15: "пятнадцать", 16: "шестнадцать", 17: "семнадцать", 18: "восемнадцать",
    19: "девятнадцать", 20: "двадцать",
}
_RU_TENS = {
    30: "тридцать", 40: "сорок", 50: "пятьдесят", 60: "шестьдесят",
    70: "семьдесят", 80: "восемьдесят", 90: "девяносто",
}
_RU_HUNDREDS = {100: "сто"}


def _num_to_ru_1_100(n: int) -> Optional[str]:
    if n in _RU_NUM_0_20:
        return _RU_NUM_0_20[n]
    if n in _RU_TENS:
        return _RU_TENS[n]
    if 21 <= n <= 29:
        return "двадцать " + _RU_NUM_0_20[n - 20]
    if 31 <= n <= 39:
        return "тридцать " + _RU_NUM_0_20[n - 30]
    if 41 <= n <= 49:
        return "сорок " + _RU_NUM_0_20[n - 40]
    if 51 <= n <= 59:
        return "пятьдесят " + _RU_NUM_0_20[n - 50]
    if 61 <= n <= 69:
        return "шестьдесят " + _RU_NUM_0_20[n - 60]
    if 71 <= n <= 79:
        return "семьдесят " + _RU_NUM_0_20[n - 70]
    if 81 <= n <= 89:
        return "восемьдесят " + _RU_NUM_0_20[n - 80]
    if 91 <= n <= 99:
        return "девяносто " + _RU_NUM_0_20[n - 90]
    if n == 100:
        return _RU_HUNDREDS[100]
    return None


def fix_digit_word_mismatch_ru(text: str) -> str:
    if not isinstance(text, str) or not text:
        return text or ""

    def _sub(m: re.Match) -> str:
        num_s = m.group(1)
        inside = m.group(2).strip()
        try:
            n = int(num_s)
        except Exception:
            return m.group(0)

        ru = _num_to_ru_1_100(n)
        if not ru:
            return m.group(0)

        if not re.search(r"[А-Яа-яЁё]", inside):
            return m.group(0)

        inside_low = inside.lower()
        if ru.split()[0] in inside_low:
            return m.group(0)

        return f"{num_s} ({ru})"

    return _RE_DIGIT_PAREN.sub(_sub, text)


def sanitize_junk_artifacts(text: str) -> str:
    if not isinstance(text, str) or not text:
        return text or ""
    # удаляем скобочные вставки с Деяния и т.п.
    s = re.sub(r"\([^)]*\bДеяни[яе][^)]*\)", "", text, flags=re.IGNORECASE)
    return _normalize_spaces(s)


def apply_post_pipeline_ru(text: str) -> str:
    s = apply_term_normalization(text)
    s = fix_digit_word_mismatch_ru(s)
    s = sanitize_junk_artifacts(s)
    return _normalize_spaces(s)


# =============================================================================
# DOCX: вариант A (абзацы)
# =============================================================================

def collect_docx_items(path: str) -> List[Dict[str, object]]:
    items: List[Dict[str, object]] = []
    with zipfile.ZipFile(path, "r") as zin:
        for info in zin.infolist():
            fname = info.filename
            if not (fname.startswith("word/") and fname.lower().endswith(".xml")):
                continue
            xml_bytes = zin.read(fname)
            try:
                root = ET.fromstring(xml_bytes)
            except Exception:
                continue

            m = re.match(r"\{(.*)\}", root.tag)
            ns = m.group(1) if m else ""
            p_tag = f"{{{ns}}}p"
            t_tag = f"{{{ns}}}t"

            for p_index, p in enumerate(root.iter(p_tag)):
                t_elems = list(p.iter(t_tag))
                if not t_elems:
                    continue
                full_text = "".join([(t.text or "") for t in t_elems])
                if not full_text:
                    continue
                if not GEORGIAN_RE.search(full_text):
                    continue
                clean_text = full_text.strip()
                if not clean_text:
                    continue
                item_id = f"{fname}::p{p_index}"
                items.append({
                    "id": item_id,
                    "xml_name": fname,
                    "p_index": p_index,
                    "full_text": full_text,
                    "clean_text": clean_text,
                })
    print(f"📄 DOCX: найдено {len(items)} абзацев с грузинским текстом.")
    return items


def process_docx_xml_paragraphs(xml_bytes: bytes, xml_name: str, id_mapping: Dict[str, str]) -> bytes:
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return xml_bytes

    m = re.match(r"\{(.*)\}", root.tag)
    ns = m.group(1) if m else ""
    p_tag = f"{{{ns}}}p"
    t_tag = f"{{{ns}}}t"

    for p_index, p in enumerate(root.iter(p_tag)):
        para_id = f"{xml_name}::p{p_index}"
        if para_id not in id_mapping:
            continue

        t_elems = list(p.iter(t_tag))
        if not t_elems:
            continue

        orig_full = "".join([(t.text or "") for t in t_elems])
        if not orig_full:
            continue
        if not GEORGIAN_RE.search(orig_full):
            continue

        translated_clean = id_mapping[para_id]
        lead = len(orig_full) - len(orig_full.lstrip())
        trail = len(orig_full) - len(orig_full.rstrip())
        prefix = orig_full[:lead]
        suffix = orig_full[len(orig_full) - trail:] if trail > 0 else ""
        translated_full = prefix + translated_clean + suffix

        t_elems[0].text = translated_full
        for t in t_elems[1:]:
            t.text = ""

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def apply_translations_docx(
    input_path: str,
    output_path: str,
    id_mapping: Dict[str, str],
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 90.0,
    end: float = 100.0,
) -> None:
    with zipfile.ZipFile(input_path, "r") as zin, \
         zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        infos = zin.infolist()
        total = len(infos) if infos else 1
        changed = 0

        for idx, info in enumerate(infos, start=1):
            fname = info.filename
            data = zin.read(fname)
            new_data = data

            if fname.startswith("word/") and fname.lower().endswith(".xml"):
                new_data = process_docx_xml_paragraphs(new_data, fname, id_mapping)

            if new_data != data:
                changed += 1

            zout.writestr(info, new_data)

            if progress_callback and total > 0:
                pct = start + (end - start) * (idx / total)
                progress_callback(pct, "Применение перевода в DOCX...")

    print(f"💾 DOCX сохранён: {output_path}, изменённых XML: {changed}")


def debug_scan_docx_for_georgian(path: str, max_examples: int = 20) -> None:
    count = 0
    examples = []
    with zipfile.ZipFile(path, "r") as zin:
        for info in zin.infolist():
            fname = info.filename
            if not (fname.startswith("word/") and fname.lower().endswith(".xml")):
                continue
            xml_bytes = zin.read(fname)
            try:
                root = ET.fromstring(xml_bytes)
            except Exception:
                continue
            m = re.match(r"\{(.*)\}", root.tag)
            ns = m.group(1) if m else ""
            p_tag = f"{{{ns}}}p"
            t_tag = f"{{{ns}}}t"
            for p_index, p in enumerate(root.iter(p_tag)):
                full_text = "".join([(t.text or "") for t in p.iter(t_tag)])
                if full_text and GEORGIAN_RE.search(full_text):
                    count += 1
                    if len(examples) < max_examples:
                        snippet = full_text.strip()
                        if len(snippet) > 120:
                            snippet = snippet[:117] + "..."
                        examples.append((fname, p_index, snippet))
    print(f"🔍 В файле {os.path.basename(path)} осталось абзацев с грузинским: {count}")
    for fname, p_index, snippet in examples:
        print(f"  - {fname}::p{p_index}: {snippet}")


# =============================================================================
# XLSX: перевод по ячейкам целиком
# =============================================================================

def collect_xlsx_cell_items(path: str) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False)

    items: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                if v.startswith("="):
                    continue
                if not GEORGIAN_RE.search(v):
                    continue
                clean = v.strip()
                if not clean:
                    continue
                item_id = f"{ws.title}::{cell.coordinate}"
                items.append({
                    "id": item_id,
                    "sheet": ws.title,
                    "coord": cell.coordinate,
                    "full_text": v,
                    "clean_text": clean,
                })

    print(f"📊 XLSX: найдено {len(items)} ячеек со строками на грузинском.")
    return items


def apply_translations_xlsx_cells(
    input_path: str,
    output_path: str,
    id_mapping: Dict[str, str],
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 90.0,
    end: float = 100.0,
) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(input_path, data_only=False)

    total = len(id_mapping) if id_mapping else 1
    done = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                item_id = f"{ws.title}::{cell.coordinate}"
                if item_id not in id_mapping:
                    continue
                v = cell.value
                if not isinstance(v, str) or v.startswith("="):
                    continue

                translated_clean = id_mapping[item_id]
                lead = len(v) - len(v.lstrip())
                trail = len(v) - len(v.rstrip())
                prefix = v[:lead]
                suffix = v[len(v) - trail:] if trail > 0 else ""
                cell.value = f"{prefix}{translated_clean}{suffix}"

                done += 1
                if progress_callback and total > 0:
                    pct = start + (end - start) * (done / total)
                    progress_callback(pct, "Применение перевода в XLSX...")

    wb.save(output_path)
    print(f"💾 XLSX сохранён: {output_path}, изменённых ячеек: {done}")


# =============================================================================
# PPTX: перевод по параграфам a:p целиком
# =============================================================================

A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"


def collect_pptx_paragraph_items(path: str) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    with zipfile.ZipFile(path, "r") as zin:
        for info in zin.infolist():
            fname = info.filename
            low = fname.lower()
            if not (low.startswith("ppt/") and low.endswith(".xml")):
                continue
            if not (
                low.startswith("ppt/slides/")
                or low.startswith("ppt/notesslides/")
                or low.startswith("ppt/slidelayouts/")
                or low.startswith("ppt/slidemasters/")
                or low == "ppt/presentation.xml"
            ):
                continue

            xml_bytes = zin.read(fname)
            try:
                root = ET.fromstring(xml_bytes)
            except Exception:
                continue

            p_tag = f"{{{A_NS}}}p"
            t_tag = f"{{{A_NS}}}t"

            p_index = 0
            for p in root.iter(p_tag):
                t_elems = list(p.iter(t_tag))
                if not t_elems:
                    continue
                full_text = "".join([(t.text or "") for t in t_elems])
                if not full_text or not GEORGIAN_RE.search(full_text):
                    continue
                clean = full_text.strip()
                if not clean:
                    continue

                items.append({
                    "id": f"{fname}::p{p_index}",
                    "xml_name": fname,
                    "p_index": p_index,
                    "full_text": full_text,
                    "clean_text": clean,
                })
                p_index += 1

    print(f"📽️ PPTX: найдено {len(items)} абзацев (a:p) с грузинским.")
    return items


def process_pptx_xml_paragraphs(xml_bytes: bytes, xml_name: str, id_mapping: Dict[str, str]) -> bytes:
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return xml_bytes

    p_tag = f"{{{A_NS}}}p"
    t_tag = f"{{{A_NS}}}t"

    p_index = 0
    for p in root.iter(p_tag):
        para_id = f"{xml_name}::p{p_index}"
        t_elems = list(p.iter(t_tag))
        if not t_elems:
            continue

        full_text = "".join([(t.text or "") for t in t_elems])
        if not full_text:
            continue

        if para_id in id_mapping and GEORGIAN_RE.search(full_text):
            translated_clean = id_mapping[para_id]
            lead = len(full_text) - len(full_text.lstrip())
            trail = len(full_text) - len(full_text.rstrip())
            prefix = full_text[:lead]
            suffix = full_text[len(full_text) - trail:] if trail > 0 else ""
            translated_full = prefix + translated_clean + suffix

            t_elems[0].text = translated_full
            for t in t_elems[1:]:
                t.text = ""

        p_index += 1

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def apply_translations_pptx_paragraphs(
    input_path: str,
    output_path: str,
    id_mapping: Dict[str, str],
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 90.0,
    end: float = 100.0,
) -> None:
    with zipfile.ZipFile(input_path, "r") as zin, \
         zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        infos = zin.infolist()
        total = len(infos) if infos else 1
        changed = 0

        for idx, info in enumerate(infos, start=1):
            fname = info.filename
            data = zin.read(fname)
            new_data = data

            low = fname.lower()
            if low.startswith("ppt/") and low.endswith(".xml"):
                new_data = process_pptx_xml_paragraphs(new_data, fname, id_mapping)

            if new_data != data:
                changed += 1

            zout.writestr(info, new_data)

            if progress_callback and total > 0:
                pct = start + (end - start) * (idx / total)
                progress_callback(pct, "Применение перевода в PPTX...")

    print(f"💾 PPTX сохранён: {output_path}, изменённых XML: {changed}")


# =============================================================================
# NLLB локальный переводчик (вариант A: абзац/параграф целиком + словарь терминов)
# =============================================================================

_PH_RE_CANON = re.compile(r"__PH\d+__")
_PH_RE_FUZZY = re.compile(r"__\s*PH\s*(\d+)\s*__")
_SENT_BOUNDARY = re.compile(r"(?<!\b\d)([.!?;:])(\s+)")


def _normalize_placeholders(s: str) -> str:
    if not isinstance(s, str) or not s:
        return s or ""
    return _PH_RE_FUZZY.sub(lambda m: f"__PH{m.group(1)}__", s)


def _placeholders_set(s: str) -> set:
    s = _normalize_placeholders(s or "")
    return set(_PH_RE_CANON.findall(s))


def _freeze_legal_entities(text: str) -> Tuple[str, Dict[str, str]]:
    """
    КРИТИЧЕСКОЕ: фиксированные юр-термины (ka->ru) под плейсхолдер, чтобы NLLB НЕ мог сорваться.
    """
    if not isinstance(text, str) or not text:
        return text, {}

    repl: Dict[str, str] = {}
    idx = 0

    # --- фиксированные термины (ka -> ru), с поддержкой грузинских кавычек „…“
    fixed_terms: List[Tuple[re.Pattern, str]] = [
        (re.compile(r'(?<!\w)([„"«]?)\s*დამკვეთი\s*([”"»]?)', re.IGNORECASE), "Заказчик"),
        (re.compile(r'(?<!\w)([„"«]?)\s*შემსრულებელი\s*([”"»]?)', re.IGNORECASE), "Исполнитель"),
        (re.compile(r'(?<!\w)([„"«]?)\s*ხელშეკრულება\s*([”"»]?)', re.IGNORECASE), "Договор"),
        (re.compile(r'(?<!\w)([„"«]?)\s*მხარეები\s*([”"»]?)', re.IGNORECASE), "Стороны"),
        (re.compile(r'(?<!\w)([„"«]?)\s*ობიექტი\s*([”"»]?)', re.IGNORECASE), "Объект"),
        (re.compile(r'(?<!\w)([„"«]?)\s*საგარანტიო\s+თანხა\s*([”"»]?)', re.IGNORECASE), "гарантийная сумма"),
        (re.compile(r'(?<!\w)([„"«]?)\s*პირგასამტეხლო\s*([”"»]?)', re.IGNORECASE), "неустойка"),
        (re.compile(r'(?<!\w)([„"«]?)\s*საურავი\s*([”"»]?)', re.IGNORECASE), "пеня"),
    ]

    def _sub_fixed(m: re.Match, replacement_ru: str) -> str:
        nonlocal idx
        left_q = m.group(1) or ""
        right_q = m.group(2) or ""
        key = f"__PH{idx}__"
        idx += 1
        repl[key] = f"{left_q}{replacement_ru}{right_q}"
        return key

    s = text
    for rx, ru_term in fixed_terms:
        s = rx.sub(lambda m, ru_term=ru_term: _sub_fixed(m, ru_term), s)

    # --- остальное (числа/даты/НДС/имена/ТЭЛМИКО/Теласи)
    patterns = [
        r"(?:\bდღგ(?:-ს|-ის)?\b)",
        r"(?:\bთელმიკო(?:-ს|-ის|-სთვის)?\b)",
        r"(?:\bთელასი(?:-ს|-ის|-სთვის)?\b)",
        r"(?:\bТЭЛМИКО\b)",
        r"(?:\bТеласи\b)",
        r"(?:(?:№|#|N)\s?\d+(?:[/-]\d+){0,3})",
        r"(?:\b(?:Art\.|Article|ст\.|Статья|п\.|пп\.|Пункт|параграф)\s*\d+(?:\.\d+){0,3}\b)",
        r"(?:\b\d+(?:\.\d+){1,4}\b)",
        r"(?:\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b)",
        r"(?:\b\d{4}[./-]\d{1,2}[./-]\d{1,2}\b)",
        r"(?:\b\d[\d\s\u00A0]*[.,]\d+\b)",
        r"(?:\b\d+(?:[.,]\d+)?\s?(?:%|GEL|USD|EUR|kWh|MWh|GWh|kV|MW|kW|ლარი|₾|\$|€)\b)",
        r"(?:\b\d{4,}\b)",
    ]

    big_pat = re.compile("|".join(f"(?:{p})" for p in patterns))

    def _sub_general(m: re.Match) -> str:
        nonlocal idx
        val = m.group(0) or ""
        val = _normalize_placeholders(val)
        if _PH_RE_CANON.fullmatch(val):
            return val
        key = f"__PH{idx}__"
        idx += 1
        repl[key] = val
        return key

    frozen = big_pat.sub(_sub_general, s)
    return frozen, repl


def _unfreeze_legal_entities(text: str, repl: Dict[str, str]) -> str:
    if not isinstance(text, str) or not repl:
        return text
    out = _normalize_placeholders(text)
    for k, v in repl.items():
        out = out.replace(k, v)
    return out


def _split_with_separators(text: str, max_len: int = 2200, soft_limit: int = 2000) -> List[Tuple[str, str]]:
    if not text:
        return [("", "")]
    if len(text) <= max_len:
        return [(text, "")]

    out: List[Tuple[str, str]] = []
    buf_start = 0
    for m in _SENT_BOUNDARY.finditer(text):
        end_punct = m.end(1)
        end_ws = m.end(2)
        if len(text[buf_start:end_ws]) < soft_limit:
            continue
        chunk = text[buf_start:end_punct]
        sep = text[end_punct:end_ws]
        out.append((chunk, sep))
        buf_start = end_ws

    if buf_start < len(text):
        out.append((text[buf_start:], ""))

    if len(out) > 40 or (len(out) == 1 and len(out[0][0]) > max_len * 2):
        out = []
        step = 1800
        i = 0
        while i < len(text):
            j = min(len(text), i + step)
            out.append((text[i:j], ""))
            i = j

    out = [(c, s) for (c, s) in out if c and c.strip()]
    if not out:
        return [(text, "")]
    return out


def _rejoin_with_separators(parts: List[Tuple[str, str]]) -> str:
    return "".join([c + s for (c, s) in parts]).strip()


def _load_nllb_tokenizer_and_model(model_name: str, src_lang: str, device: torch.device) -> Tuple[Any, Any]:
    hf_enable_offline()
    try:
        tokenizer = AutoTokenizer.from_pretrained(model_name, src_lang=src_lang, local_files_only=True)
        model = AutoModelForSeq2SeqLM.from_pretrained(
            model_name,
            local_files_only=True,
            torch_dtype=torch.float16 if device.type == "cuda" else None,
        ).to(device)
        return tokenizer, model
    except Exception:
        print("⚠️ NLLB не найден локально. Включаю ONLINE и скачиваю модель...")
        hf_enable_online()
        tokenizer = AutoTokenizer.from_pretrained(model_name, src_lang=src_lang, local_files_only=False)
        model = AutoModelForSeq2SeqLM.from_pretrained(
            model_name,
            local_files_only=False,
            torch_dtype=torch.float16 if device.type == "cuda" else None,
        ).to(device)
        hf_enable_offline()
        print("✅ NLLB скачан. Возвращаю OFFLINE.")
        return tokenizer, model


def translate_with_local_model(
    fragments: List[str],
    direction_code: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 10.0,
    end: float = 60.0,
) -> Dict[str, str]:
    LANG_MAP = {"ka": "kat_Geor", "ru": "rus_Cyrl", "en": "eng_Latn"}

    if "-" not in direction_code:
        raise ValueError(f"direction_code должен быть формата ka-ru, а получено: {direction_code}")

    src, tgt = direction_code.split("-")
    if src not in LANG_MAP:
        raise ValueError(f"Источник языка '{src}' не поддержан")
    if tgt not in LANG_MAP:
        raise ValueError(f"Целевой язык '{tgt}' не поддержан")

    SRC_LANG = LANG_MAP[src]
    TGT_LANG = LANG_MAP[tgt]

    remaining = [f.strip() for f in fragments if isinstance(f, str) and f.strip()]
    if not remaining:
        return {}

    MODEL_NAME = "facebook/nllb-200-3.3B"
    if progress_callback:
        progress_callback(start, f"Загрузка локальной модели NLLB ({MODEL_NAME})…")
    print(f"⏳ Загружаем локальную NLLB модель: {MODEL_NAME}")
    hf_print_mode("[Before load] ")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    tokenizer, model = _load_nllb_tokenizer_and_model(MODEL_NAME, SRC_LANG, device)
    model.eval()

    if hasattr(tokenizer, "lang_code_to_id") and TGT_LANG in tokenizer.lang_code_to_id:
        forced_bos_id = tokenizer.lang_code_to_id[TGT_LANG]
    else:
        forced_bos_id = tokenizer.convert_tokens_to_ids(TGT_LANG)
    if forced_bos_id is None or forced_bos_id < 0:
        raise RuntimeError(f"Не удалось получить token id для языка {TGT_LANG}")

    gen_kwargs = dict(
        forced_bos_token_id=forced_bos_id,
        do_sample=False,
        num_beams=8,
        length_penalty=1.1,
        no_repeat_ngram_size=3,
        repetition_penalty=1.05,
        early_stopping=True,
        use_cache=True,
        max_new_tokens=1024,
        pad_token_id=tokenizer.pad_token_id,
        eos_token_id=tokenizer.eos_token_id,
    )

    BATCH_SIZE = 4 if device.type == "cuda" else 1

    def _translate_texts(texts: List[str]) -> List[str]:
        tokenizer.src_lang = SRC_LANG
        inputs = tokenizer(texts, return_tensors="pt", padding=True, truncation=False)
        inputs = {k: v.to(device) for k, v in inputs.items()}
        with torch.no_grad():
            generated = model.generate(**inputs, **gen_kwargs)
        outs = tokenizer.batch_decode(generated, skip_special_tokens=True)
        return [_normalize_placeholders((o or "").strip()) for o in outs]

    total = len(remaining)
    done = 0
    mapping: Dict[str, str] = {}

    for batch in chunks(remaining, BATCH_SIZE):
        batch_parts: List[List[Tuple[str, str]]] = []
        batch_frozen: List[List[str]] = []
        batch_meta: List[List[Dict[str, str]]] = []
        batch_raw: List[List[str]] = []

        for orig in batch:
            parts = _split_with_separators(orig)
            batch_parts.append(parts)

            frozen_chunks: List[str] = []
            metas: List[Dict[str, str]] = []
            raw_chunks: List[str] = []

            for (chunk_text, _sep) in parts:
                raw_chunks.append(chunk_text)
                fp, meta = _freeze_legal_entities(chunk_text)
                frozen_chunks.append(_normalize_placeholders(fp))
                metas.append(meta)

            batch_frozen.append(frozen_chunks)
            batch_meta.append(metas)
            batch_raw.append(raw_chunks)

        flat_in = [c for row in batch_frozen for c in row]
        flat_out = _translate_texts(flat_in)

        cursor = 0
        for orig, parts, frozen_chunks, metas, raw_chunks in zip(batch, batch_parts, batch_frozen, batch_meta, batch_raw):
            n = len(frozen_chunks)
            out_chunks = flat_out[cursor:cursor + n]
            cursor += n

            rebuilt: List[Tuple[str, str]] = []
            for (chunk_text, sep), fp, out_text, meta, raw_p in zip(parts, frozen_chunks, out_chunks, metas, raw_chunks):
                fp_norm = _normalize_placeholders(fp)
                t = _normalize_placeholders((out_text or "").strip())

                if not t:
                    retry = _translate_texts([fp_norm])[0] if fp_norm else ""
                    t = retry or ""

                # строгая проверка плейсхолдеров
                ph_before = _placeholders_set(fp_norm)
                ph_after = _placeholders_set(t)
                if ph_before and (ph_before != ph_after):
                    retry = _translate_texts([fp_norm])[0] if fp_norm else ""
                    if retry:
                        t = retry
                        ph_after = _placeholders_set(t)

                if ph_before and (ph_before != ph_after):
                    retry_raw = _translate_texts([raw_p])[0] if raw_p else ""
                    if retry_raw:
                        t = retry_raw

                t = _unfreeze_legal_entities(t, meta)
                t = _normalize_spaces(t)

                if tgt == "ru":
                    t = apply_post_pipeline_ru(t)

                # если остался грузинский — fallback на перевод raw
                if src == "ka" and tgt != "ka" and _has_ka(t):
                    repaired = _translate_texts([raw_p])[0] if raw_p else ""
                    if repaired and not _has_ka(repaired):
                        repaired = _normalize_spaces(repaired)
                        if tgt == "ru":
                            repaired = apply_post_pipeline_ru(repaired)
                        t = repaired

                rebuilt.append((t if t else chunk_text, sep))

            merged = _rejoin_with_separators(rebuilt)
            merged = _normalize_spaces(merged)
            if tgt == "ru":
                merged = apply_post_pipeline_ru(merged)

            mapping[orig] = merged if merged else orig

        done += len(batch)
        if progress_callback and total > 0:
            pct = start + (end - start) * (done / total)
            progress_callback(pct, f"Перевод локальной моделью (NLLB, {direction_code})…")
        print(f"   [NLLB {direction_code}] готово {done}/{total}")

    return mapping


# =============================================================================
# Qwen post-edit (опционально)
# =============================================================================

def _load_qwen_tokenizer_and_model(model_name: str, device: torch.device):
    hf_enable_offline()
    try:
        tokenizer = AutoTokenizer.from_pretrained(model_name, trust_remote_code=True, local_files_only=True)
        if device.type == "cuda":
            model = AutoModelForCausalLM.from_pretrained(
                model_name,
                torch_dtype=torch.float16,
                device_map="auto",
                trust_remote_code=True,
                local_files_only=True,
            )
        else:
            model = AutoModelForCausalLM.from_pretrained(
                model_name,
                trust_remote_code=True,
                local_files_only=True,
            ).to(device)
        return tokenizer, model
    except Exception:
        print("⚠️ Qwen не найден локально. Включаю ONLINE и скачиваю модель...")
        hf_enable_online()
        tokenizer = AutoTokenizer.from_pretrained(model_name, trust_remote_code=True, local_files_only=False)
        if device.type == "cuda":
            model = AutoModelForCausalLM.from_pretrained(
                model_name,
                torch_dtype=torch.float16,
                device_map="auto",
                trust_remote_code=True,
                local_files_only=False,
            )
        else:
            model = AutoModelForCausalLM.from_pretrained(
                model_name,
                trust_remote_code=True,
                local_files_only=False,
            ).to(device)
        hf_enable_offline()
        print("✅ Qwen скачан. Возвращаю OFFLINE.")
        return tokenizer, model


def qwen_postedit_prompt_ru() -> str:
    return (
        "You are a professional editor for Russian legal contracts and official documents.\n"
        "Task: improve legal style, clarity, grammar and fluency.\n"
        "Rules:\n"
        "1) Preserve meaning EXACTLY. Do not invent content.\n"
        "2) Preserve ALL numbers, dates, names, bank details.\n"
        "3) Keep proper names unchanged: 'ТЭЛМИКО' and 'Теласи' (do NOT decline).\n"
        "4) Fix MT artifacts: 'единый счет казино' -> 'единый казначейский счет', "
        "'пергасаточка/пергазмаплата' -> 'неустойка', digit+(words) must match.\n"
        "5) Output ONLY the improved text. No explanations. No markdown.\n"
    )


def _qwen_postedit_should_run(text: str) -> bool:
    if not isinstance(text, str) or not text.strip():
        return False
    if _has_ka(text):
        return True
    low = text.lower()
    bad = ["казино", "пергасат", "пергазмаплат", "департамент обороны", "r/n", "стороne"]
    if any(x in low for x in bad):
        return True
    if _RE_JUNK_BIBLE.search(text):
        return True
    if re.search(r"\b\d+\s*\(", text):
        return True
    return False


def _looks_like_qwen_junk(out_text: str, src_len: int) -> bool:
    low = (out_text or "").lower()
    if "```" in out_text:
        return True
    if any(x in low for x in ["объяснен", "комментар", "вот исправ", "правки:", "json"]):
        return True
    if src_len > 0 and len(out_text) > int(src_len * 1.8):
        return True
    return False


def post_edit_with_qwen_local(
    mapping: Dict[str, str],
    target_language: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 60.0,
    end: float = 90.0,
) -> Dict[str, str]:
    MODEL_NAME = "Qwen/Qwen2.5-3B-Instruct"

    to_edit: List[str] = []
    for v in mapping.values():
        if isinstance(v, str) and v.strip() and _qwen_postedit_should_run(v):
            to_edit.append(v)

    unique_values: List[str] = []
    seen = set()
    for v in to_edit:
        if v not in seen:
            seen.add(v)
            unique_values.append(v)

    if not unique_values:
        return mapping

    print(f"⏳ Загружаем локальную модель вычитки: {MODEL_NAME}")
    if progress_callback:
        progress_callback(start, "Загрузка Qwen2.5-3B-Instruct для вычитки…")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    tokenizer, model = _load_qwen_tokenizer_and_model(MODEL_NAME, device)
    model.eval()

    system_msg_text = qwen_postedit_prompt_ru() if target_language.lower() == "russian" else (
        f"You are a professional editor for {target_language} legal documents. Output ONLY improved text."
    )

    total = len(unique_values)
    done = 0
    improved_map: Dict[str, str] = {}

    for text in unique_values:
        pre = apply_post_pipeline_ru(text) if target_language.lower() == "russian" else text

        fp, meta = _freeze_legal_entities(pre)
        fp = _normalize_placeholders(fp)
        ph_before = _placeholders_set(fp)

        messages = [
            {"role": "system", "content": system_msg_text},
            {"role": "user", "content": fp},
        ]
        prompt = tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
        inputs = tokenizer(prompt, return_tensors="pt").to(device)

        with torch.no_grad():
            output_ids = model.generate(
                **inputs,
                max_new_tokens=900,
                do_sample=False,
            )

        gen_ids = output_ids[0][inputs["input_ids"].shape[1]:]
        out_text = tokenizer.decode(gen_ids, skip_special_tokens=True).strip()
        out_text = _normalize_placeholders(out_text)

        if _looks_like_qwen_junk(out_text, src_len=len(fp)):
            improved = pre
        else:
            ph_after = _placeholders_set(out_text)
            if ph_before and (ph_before != ph_after):
                improved = pre
            else:
                improved = _unfreeze_legal_entities(out_text, meta)
                improved = _normalize_spaces(improved)

        if target_language.lower() == "russian":
            improved = apply_post_pipeline_ru(improved)

        improved_map[text] = improved if improved else text

        done += 1
        if progress_callback and total > 0:
            pct = start + (end - start) * (done / total)
            progress_callback(pct, "Литературная вычитка (локальная Qwen)…")

    new_mapping: Dict[str, str] = {}
    for geo, raw_trans in mapping.items():
        if isinstance(raw_trans, str) and raw_trans in improved_map:
            new_mapping[geo] = improved_map[raw_trans]
        else:
            new_mapping[geo] = raw_trans

    return new_mapping


# =============================================================================
# ChatGPT перевод + постредактура (сохранено как было)
# =============================================================================

def translate_with_chatgpt(
    fragments: List[str],
    model_name: str,
    api_key: str,
    target_language: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 10.0,
    end: float = 60.0,
) -> Dict[str, str]:
    from openai import OpenAI

    os.environ["OPENAI_API_KEY"] = api_key
    client = OpenAI()

    cleaned = [s.strip() for s in fragments if isinstance(s, str) and s.strip()]
    unique_texts: List[str] = []
    seen: Set[str] = set()
    for s in cleaned:
        if s not in seen:
            seen.add(s)
            unique_texts.append(s)

    if not unique_texts:
        return {}

    id_to_text = {i: txt for i, txt in enumerate(unique_texts)}
    text_to_id = {txt: i for i, txt in id_to_text.items()}

    fragments_struct = [{"id": i, "text": txt} for i, txt in id_to_text.items()]
    batches = list(split_fragments_by_tokens(fragments_struct, max_tokens_per_batch=8000))

    total = len(unique_texts)
    done = 0
    id_to_translated: Dict[int, str] = {}

    system_msg = (
        "You are a professional legal and technical translator.\n"
        f"Translate from Georgian to {target_language} into natural, formal, human-quality language.\n"
        "Preserve all facts, numbers, names, codes, bank details.\n"
        "Proper names: TELMICO and Telasi must be kept as proper names. "
        "In Russian use exactly 'ТЭЛМИКО' and 'Теласи' and NEVER decline them.\n"
        "Return ONLY JSON: {\"translations\": [{\"id\":..., \"text\":...}, ...]}.\n"
    )

    for batch_idx, batch in enumerate(batches, start=1):
        if progress_callback and total > 0:
            pct = start + (end - start) * (done / total)
            progress_callback(pct, "Перевод через ChatGPT...")

        user_payload = {"source_language": "Georgian", "target_language": target_language, "items": batch}

        max_retries = 5
        delay_seconds = 10

        for attempt in range(1, max_retries + 1):
            try:
                resp = client.chat.completions.create(
                    model=model_name,
                    response_format={"type": "json_object"},
                    messages=[
                        {"role": "system", "content": system_msg},
                        {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
                    ],
                )
                break
            except RateLimitError:
                if attempt == max_retries:
                    raise
                time.sleep(delay_seconds)

        data = json.loads(resp.choices[0].message.content)
        translations_list = data.get("translations")
        if not isinstance(translations_list, list):
            raise ValueError("Ожидался ключ 'translations' со списком объектов {id, text}.")

        for obj in translations_list:
            if not isinstance(obj, dict):
                continue
            tid = obj.get("id")
            ttext = obj.get("text")
            try:
                tid_int = int(tid)
            except (TypeError, ValueError):
                continue
            if isinstance(ttext, str) and ttext.strip():
                id_to_translated[tid_int] = ttext.strip()

        done += len(batch)
        print(f"   ChatGPT перевёл {done}/{total} (batch {batch_idx}/{len(batches)})")

    mapping: Dict[str, str] = {}
    for txt, tid in text_to_id.items():
        out = id_to_translated.get(tid, txt)
        if target_language.lower() == "russian":
            out = apply_post_pipeline_ru(out)
        mapping[txt] = out

    return mapping


def post_edit_with_chatgpt(
    mapping: Dict[str, str],
    model_name: str,
    api_key: str,
    target_language: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
    start: float = 70.0,
    end: float = 90.0,
) -> Dict[str, str]:
    from openai import OpenAI

    os.environ["OPENAI_API_KEY"] = api_key
    client = OpenAI()

    unique_values: List[str] = []
    seen = set()
    for v in mapping.values():
        if isinstance(v, str) and v.strip() and v not in seen:
            seen.add(v)
            unique_values.append(v)

    if not unique_values:
        return mapping

    batches = list(split_texts_by_tokens(unique_values, max_tokens_per_batch=8000))
    total = len(unique_values)
    done = 0

    system_msg = (
        "You are a professional editor for legal contracts and official documents.\n"
        f"Improve style and clarity in {target_language} while preserving meaning and facts.\n"
        "Do NOT invent content. Keep numbers and bank details unchanged.\n"
        "Keep proper names unchanged: Russian 'ТЭЛМИКО' and 'Теласи' (never decline).\n"
        "Fix MT artifacts: casino->treasury, nonsense legal terms->standard, digit+(words) must match.\n"
        "Return ONLY JSON mapping: {\"<original>\": \"<improved>\", ...}.\n"
    )

    improved_map: Dict[str, str] = {}

    for batch_idx, batch in enumerate(batches, start=1):
        if progress_callback and total > 0:
            pct = start + (end - start) * (done / total)
            progress_callback(pct, "Литературная вычитка (ChatGPT)...")

        user_payload = {"target_language": target_language, "texts": batch}

        max_retries = 5
        delay_seconds = 10
        for attempt in range(1, max_retries + 1):
            try:
                resp = client.chat.completions.create(
                    model=model_name,
                    response_format={"type": "json_object"},
                    messages=[
                        {"role": "system", "content": system_msg},
                        {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
                    ],
                )
                break
            except RateLimitError:
                if attempt == max_retries:
                    raise
                time.sleep(delay_seconds)

        data = json.loads(resp.choices[0].message.content)
        for orig_text in batch:
            new_text = data.get(orig_text)
            out = new_text.strip() if isinstance(new_text, str) and new_text.strip() else orig_text
            if target_language.lower() == "russian":
                out = apply_post_pipeline_ru(out)
            improved_map[orig_text] = out

        done += len(batch)
        print(f"   ChatGPT вычитал {done}/{total} (batch {batch_idx}/{len(batches)})")

    new_mapping: Dict[str, str] = {}
    for geo, raw_trans in mapping.items():
        new_mapping[geo] = improved_map.get(raw_trans, raw_trans) if isinstance(raw_trans, str) else raw_trans
    return new_mapping


# =============================================================================
# Основной pipeline для DOCX/XLSX/PPTX
# =============================================================================

def process_file(
    file_path: str,
    translator_kind: str,
    chatgpt_model: str,
    env_path: Optional[str],
    direction_code: str,
    post_edit: bool,
    progress_callback: Optional[Callable[[float, str], None]] = None,
) -> str:
    if progress_callback is None:
        def progress_callback(pct: float, msg: str) -> None:
            pass

    if not (is_docx(file_path) or is_xlsx(file_path) or is_pptx(file_path)):
        raise ValueError("Поддерживаются только файлы .docx, .xlsx и .pptx")

    if direction_code not in DIRECTION_CONFIG:
        raise ValueError(f"Неизвестное направление: {direction_code}")

    meta = DIRECTION_CONFIG[direction_code]
    target_language = meta["target_language"]
    suffix = meta["suffix"]

    progress_callback(0.0, "Сбор грузинских фрагментов...")

    docx_items = None
    xlsx_items = None
    pptx_items = None

    if is_docx(file_path):
        docx_items = collect_docx_items(file_path)
        if not docx_items:
            raise RuntimeError("В DOCX не найден грузинский текст для перевода.")
        fragments_for_translation = sorted({str(it["clean_text"]) for it in docx_items})

    elif is_xlsx(file_path):
        xlsx_items = collect_xlsx_cell_items(file_path)
        if not xlsx_items:
            raise RuntimeError("В XLSX не найден грузинский текст для перевода.")
        fragments_for_translation = sorted({str(it["clean_text"]) for it in xlsx_items})

    else:
        pptx_items = collect_pptx_paragraph_items(file_path)
        if not pptx_items:
            raise RuntimeError("В PPTX не найден грузинский текст для перевода.")
        fragments_for_translation = sorted({str(it["clean_text"]) for it in pptx_items})

    print(f"Найдено {len(fragments_for_translation)} уникальных фрагментов для перевода.")
    progress_callback(5.0, f"Найдено {len(fragments_for_translation)} фрагментов. Подготовка к переводу...")

    if translator_kind == "chatgpt":
        if not env_path:
            raise ValueError("Не выбран .env файл с токеном для ChatGPT.")
        api_key = load_api_key_from_env_file(env_path)

        mapping_text_to_trans = translate_with_chatgpt(
            fragments_for_translation,
            chatgpt_model,
            api_key,
            target_language,
            progress_callback=progress_callback,
            start=10.0,
            end=60.0,
        )

        if post_edit:
            mapping_text_to_trans = post_edit_with_chatgpt(
                mapping_text_to_trans,
                chatgpt_model,
                api_key,
                target_language,
                progress_callback=progress_callback,
                start=60.0,
                end=90.0,
            )
    else:
        mapping_text_to_trans = translate_with_local_model(
            fragments_for_translation,
            direction_code,
            progress_callback=progress_callback,
            start=10.0,
            end=60.0,
        )

        # финальные фиксы после NLLB
        if target_language.lower() == "russian":
            mapping_text_to_trans = {k: apply_post_pipeline_ru(v) for k, v in mapping_text_to_trans.items()}

        if post_edit:
            mapping_text_to_trans = post_edit_with_qwen_local(
                mapping_text_to_trans,
                target_language=target_language,
                progress_callback=progress_callback,
                start=60.0,
                end=90.0,
            )
            if target_language.lower() == "russian":
                mapping_text_to_trans = {k: apply_post_pipeline_ru(v) for k, v in mapping_text_to_trans.items()}

    base, ext = os.path.splitext(file_path)
    output_path = f"{base}{suffix}{ext}"
    progress_callback(90.0, "Применяем переводы к файлу...")

    if is_docx(file_path):
        assert docx_items is not None
        id_mapping: Dict[str, str] = {}
        for it in docx_items:
            clean = str(it["clean_text"])
            item_id = str(it["id"])
            translated = mapping_text_to_trans.get(clean, clean)
            id_mapping[item_id] = translated
        apply_translations_docx(file_path, output_path, id_mapping, progress_callback, 90.0, 100.0)
        debug_scan_docx_for_georgian(output_path)

    elif is_xlsx(file_path):
        assert xlsx_items is not None
        id_mapping: Dict[str, str] = {}
        for it in xlsx_items:
            clean = str(it["clean_text"])
            item_id = str(it["id"])
            translated = mapping_text_to_trans.get(clean, clean)
            if target_language.lower() == "russian":
                translated = apply_post_pipeline_ru(translated)
            id_mapping[item_id] = translated
        apply_translations_xlsx_cells(file_path, output_path, id_mapping, progress_callback, 90.0, 100.0)

    else:
        assert pptx_items is not None
        id_mapping: Dict[str, str] = {}
        for it in pptx_items:
            clean = str(it["clean_text"])
            item_id = str(it["id"])
            translated = mapping_text_to_trans.get(clean, clean)
            if target_language.lower() == "russian":
                translated = apply_post_pipeline_ru(translated)
            id_mapping[item_id] = translated
        apply_translations_pptx_paragraphs(file_path, output_path, id_mapping, progress_callback, 90.0, 100.0)

    progress_callback(100.0, "Готово.")
    return output_path


# =============================================================================
# GUI (Tkinter)
# =============================================================================

class TranslatorGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Перевод грузинского текста в DOCX/XLSX/PPTX")

        self.file_path_var = tk.StringVar()
        self.env_path_var = tk.StringVar()
        self.translator_var = tk.StringVar(value="chatgpt")
        self.model_var = tk.StringVar(value=DEFAULT_CHATGPT_MODEL)
        self.direction_label_var = tk.StringVar(value=DIRECTION_CONFIG["ka-ru"]["label"])

        self.progress_var = tk.DoubleVar(value=0.0)
        self.status_var = tk.StringVar(value="Готов к работе.")
        self.post_edit_var = tk.BooleanVar(value=False)

        self.start_button: Optional[ttk.Button] = None
        self.post_edit_check: Optional[ttk.Checkbutton] = None
        self.direction_combo: Optional[ttk.Combobox] = None
        self.model_combo: Optional[ttk.Combobox] = None
        self.env_entry: Optional[ttk.Entry] = None
        self.env_button: Optional[ttk.Button] = None
        self.progress_bar: Optional[ttk.Progressbar] = None
        self.status_label: Optional[ttk.Label] = None

        self.build_ui()

    def build_ui(self):
        pad = 5
        frm = ttk.Frame(self.root, padding=10)
        frm.grid(row=0, column=0, sticky="nsew")

        ttk.Label(frm, text="Файл DOCX/XLSX/PPTX:").grid(row=0, column=0, sticky="w", pady=pad)
        entry_file = ttk.Entry(frm, textvariable=self.file_path_var, width=60)
        entry_file.grid(row=0, column=1, sticky="we", pady=pad)
        ttk.Button(frm, text="Выбрать...", command=self.choose_file).grid(row=0, column=2, padx=pad, pady=pad)

        ttk.Label(frm, text="Переводчик:").grid(row=1, column=0, sticky="w", pady=pad)
        ttk.Radiobutton(frm, text="ChatGPT (облачный)", variable=self.translator_var, value="chatgpt",
                        command=self.on_translator_change).grid(row=1, column=1, sticky="w", pady=pad)
        ttk.Radiobutton(frm, text="Локальная модель (NLLB-200)", variable=self.translator_var, value="local",
                        command=self.on_translator_change).grid(row=2, column=1, sticky="w", pady=pad)

        ttk.Label(frm, text="Направление перевода:").grid(row=3, column=0, sticky="w", pady=pad)
        direction_values = [meta["label"] for meta in DIRECTION_CONFIG.values()]
        self.direction_combo = ttk.Combobox(frm, textvariable=self.direction_label_var,
                                            values=direction_values, state="readonly", width=35)
        self.direction_combo.grid(row=3, column=1, sticky="w", pady=pad)

        ttk.Label(frm, text="Модель ChatGPT:").grid(row=4, column=0, sticky="w", pady=pad)
        self.model_combo = ttk.Combobox(frm, textvariable=self.model_var,
                                        values=CHATGPT_MODELS, state="readonly", width=35)
        self.model_combo.grid(row=4, column=1, sticky="w", pady=pad)

        ttk.Label(frm, text=".env с токеном:").grid(row=5, column=0, sticky="w", pady=pad)
        self.env_entry = ttk.Entry(frm, textvariable=self.env_path_var, width=60)
        self.env_entry.grid(row=5, column=1, sticky="we", pady=pad)
        self.env_button = ttk.Button(frm, text="Выбрать .env...", command=self.choose_env_file)
        self.env_button.grid(row=5, column=2, padx=pad, pady=pad)

        self.post_edit_check = ttk.Checkbutton(
            frm,
            text="Постредактура (исправлять MT-артефакты, стиль)",
            variable=self.post_edit_var,
        )
        self.post_edit_check.grid(row=6, column=0, columnspan=3, sticky="w", pady=pad)

        ttk.Label(frm, text="Прогресс:").grid(row=7, column=0, sticky="w", pady=pad)
        self.progress_bar = ttk.Progressbar(frm, maximum=100.0, variable=self.progress_var,
                                            mode="determinate", length=300)
        self.progress_bar.grid(row=7, column=1, columnspan=2, sticky="we", pady=pad)

        self.status_label = ttk.Label(frm, textvariable=self.status_var)
        self.status_label.grid(row=8, column=0, columnspan=3, sticky="w", pady=pad)

        self.start_button = ttk.Button(frm, text="Старт перевода", command=self.run_translation)
        self.start_button.grid(row=9, column=0, columnspan=3, pady=10)

        self.root.columnconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)

        self.on_translator_change()

    def _update_progress_mainthread(self, pct: float, msg: str) -> None:
        self.progress_var.set(max(0.0, min(100.0, pct)))
        self.status_var.set(msg)

    def set_progress(self, pct: float, msg: str) -> None:
        self.root.after(0, self._update_progress_mainthread, pct, msg)

    def choose_file(self):
        path = filedialog.askopenfilename(
            title="Выберите DOCX/XLSX/PPTX",
            filetypes=[("Office files", "*.docx *.xlsx *.pptx"), ("Все файлы", "*.*")],
        )
        if path:
            self.file_path_var.set(path)

    def choose_env_file(self):
        path = filedialog.askopenfilename(
            title="Выберите .env файл с OPENAI_API_KEY",
            filetypes=[("ENV files", "*.env;*.txt;*.*"), ("Все файлы", "*.*")],
        )
        if path:
            self.env_path_var.set(path)

    def on_translator_change(self):
        kind = self.translator_var.get()
        if kind == "chatgpt":
            if self.model_combo is not None:
                self.model_combo.configure(state="readonly")
            if self.env_entry is not None:
                self.env_entry.configure(state="normal")
            if self.env_button is not None:
                self.env_button.configure(state="normal")
            if self.post_edit_check is not None:
                self.post_edit_check.configure(text="Постредактура (через ChatGPT)", state="normal")
        else:
            if self.model_combo is not None:
                self.model_combo.configure(state="disabled")
            if self.env_entry is not None:
                self.env_entry.configure(state="disabled")
            if self.env_button is not None:
                self.env_button.configure(state="disabled")
            if self.post_edit_check is not None:
                self.post_edit_check.configure(text="Постредактура (локальная Qwen2.5-3B-Instruct)", state="normal")

    def run_translation(self):
        file_path = self.file_path_var.get().strip()
        if not file_path:
            messagebox.showerror("Ошибка", "Выберите файл DOCX/XLSX/PPTX.")
            return

        if not (is_docx(file_path) or is_xlsx(file_path) or is_pptx(file_path)):
            messagebox.showerror("Ошибка", "Поддерживаются только файлы .docx, .xlsx и .pptx.")
            return

        translator_kind = self.translator_var.get()
        chatgpt_model = self.model_var.get()
        env_path = self.env_path_var.get().strip() if translator_kind == "chatgpt" else None
        post_edit = bool(self.post_edit_var.get())

        direction_label = self.direction_label_var.get()
        try:
            direction_code = get_direction_code_from_label(direction_label)
        except ValueError as e:
            messagebox.showerror("Ошибка", str(e))
            return

        if translator_kind == "chatgpt" and not env_path:
            messagebox.showerror("Ошибка", "Выберите .env файл с токеном для ChatGPT.")
            return

        if self.start_button is not None:
            self.start_button.configure(state="disabled")
        self.set_progress(0.0, "Начало обработки...")

        t = threading.Thread(
            target=self._worker_translate,
            args=(file_path, translator_kind, chatgpt_model, env_path, direction_code, post_edit),
            daemon=True,
        )
        t.start()

    def _worker_translate(
        self,
        file_path: str,
        translator_kind: str,
        chatgpt_model: str,
        env_path: Optional[str],
        direction_code: str,
        post_edit: bool,
    ):
        try:
            output_path = process_file(
                file_path=file_path,
                translator_kind=translator_kind,
                chatgpt_model=chatgpt_model,
                env_path=env_path,
                direction_code=direction_code,
                post_edit=post_edit,
                progress_callback=self.set_progress,
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            err_msg = f"{type(e).__name__}: {e}"

            def show_error(msg=err_msg):
                if self.start_button is not None:
                    self.start_button.configure(state="normal")
                self._update_progress_mainthread(0.0, "Ошибка.")
                messagebox.showerror("Ошибка", f"Перевод завершился с ошибкой:\n{msg}")

            self.root.after(0, show_error)
            return

        def on_done():
            if self.start_button is not None:
                self.start_button.configure(state="normal")
            self._update_progress_mainthread(100.0, "Готово.")
            messagebox.showinfo("Готово", f"Файл переведён:\n{output_path}")

        self.root.after(0, on_done)


def main():
    root = tk.Tk()
    TranslatorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
